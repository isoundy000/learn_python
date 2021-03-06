#!/usr/bin/env python
# -*- coding:utf-8 -*-
"""
Created by lichen on 17/2/8.
执行该脚本，将会生成newfish.xlsm对应的json配置文件
"""

import time
from openpyxl import load_workbook
import os, sys, json, copy
import collections
import re
import csv
import traceback
from multiprocessing import Queue, Process, cpu_count, freeze_support

ServerPath = ""
ClientIdMap = None


def fish_config():
    """
    鱼的配置
    :return:
    """
    print "fish_config, start"
    outPath = getOutPath("fish")
    ws = getWorkBook().get_sheet_by_name("Fish")
    config = collections.OrderedDict()
    startRowNum = 4
    i = 0
    reload(sys)
    sys.setdefaultencoding("utf8")
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if not cols[3]:                 # 鱼种ID
            continue
        one = collections.OrderedDict()
        if str(cols[0]) in config:
            raise KeyError("fishId %d repeat" % int(cols[0]))
        config[str(cols[3])] = one          # {11001: {}}
        one["fishType"] = int(cols[3])
        one["name"] = unicode(cols[1])      # 鱼的名字
        one["type"] = int(cols[4])          # 鱼的类别
        if is_number(cols[5]):
            one["itemId"] = int(cols[5])    # 捕鱼掉道具 数字
        else:
            one["itemId"] = json.loads(cols[5])     # [10000,10002]
        one["score"] = int(cols[6])                 # 分值
        one["probb1"] = float("%.1f" % cols[7])     # 无HP时概率基数
        one["probb2"] = float("%.1f" % cols[8])
        one["HP"] = int(cols[9])                   # 生命
        one["value"] = int(cols[10])                # 实际价值
        if cols[12]:
            one["weaponId"] = int(cols[12])         # 武器Id
        one["prizeWheelValue"] = float(cols[13])    # 积攒价值
        one["triggerRate"] = int(cols[14])          # 触发几率(10000)
    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "fish_config, end"


def is_number(s):
    try:
        float(s)                        # 是float类型
        return True
    except ValueError:
        pass

    try:
        import unicodedata
        unicodedata.numeric(s)          # 把一个表示数字的字符串转换为浮点数返回
        return True
    except (TypeError, ValueError):
        pass

    return False


def weapon_config():
    """武器配置"""
    print "weapon_config, start"
    outPath = getOutPath("weapon")
    ws = getWorkBook().get_sheet_by_name("Weapon")
    weaponConfig = collections.OrderedDict()
    startRowNum = 4
    i = 0
    for row in ws.rows:
        i = i + 1
        cols = []
        if i < startRowNum:
            continue
        for cell in row:
            cols.append(cell.value)
        if not cols[2]:
            continue
        oneWeapon = {}
        if str(cols[0]) in weaponConfig:
            raise KeyError("weaponId %d repeat" % int(cols[0]))
        weaponConfig[str(cols[2])] = oneWeapon
        oneWeapon["weaponId"] = int(cols[2])
        oneWeapon["costBullet"] = cols[3]
        oneWeapon["power"] = cols[4]
        oneWeapon["matchAddition"] = cols[5]
        oneWeapon["wpRatio"] = cols[12]

    result = json.dumps(weaponConfig, indent=4)
        
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()
    print "weapon_config, end"


def skill_grade_config():
    print "skill_grade_config, start"
    outPath = getOutPath("skillGrade")
    ws = getWorkBook().get_sheet_by_name("SkillGrade")
    config = collections.OrderedDict()
    startRowNum = 4
    i = 0
    skillId = collections.OrderedDict()
    for row in ws.rows:
        i = i+1
        cols = []
        if i < startRowNum:
            continue
        if row[1].value not in config.keys():
            skillId = collections.OrderedDict()
            config[row[1].value] = skillId
        one = collections.OrderedDict()
        skillId[str(row[2].value)] = one
        for cell in row:
            cols.append(cell.value)
        if not cols[2]:
            continue
        one["cost"] = int(cols[3])
        one["power"] = int(cols[4])
        one["HP"] = int(cols[5])      
        one["impale"] = int(cols[6])
        one["clip"] = int(cols[7])
        one["interval"] = float(cols[8])
        one["duration"] = float(cols[9])
        one["coolDown"] = int(cols[10])
        one["weaponId"] = int(cols[11])
        one["double"] = json.loads(str(cols[12]))
        one["isReturn"] = int(cols[13])
    result = json.dumps(config, indent=4)        
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()
    print "skill_grade_config, end"


def skill_star_config():
    print "skill_star_config, start"
    outPath = getOutPath("skillStar")
    ws = getWorkBook().get_sheet_by_name("SkillStar")
    config = collections.OrderedDict()
    startRowNum = 4
    i = 0
    skillId = collections.OrderedDict()
    for row in ws.rows:
        i = i+1
        cols = []
        if i < startRowNum:
            continue
        if row[1].value not in config.keys():
            skillId = collections.OrderedDict()
            config[row[1].value] = skillId
        one = collections.OrderedDict()
        skillId[str(row[2].value)] = one
        for cell in row:
            cols.append(cell.value)
        if not cols[2]:
            continue
        abilities = []
        one["abilities"] = abilities
        for x in xrange(3, len(cols), 3):
            if not cols[x]:
                continue
            ability = collections.OrderedDict()
            ability["ability"] = int(cols[x])
            ability["valueType"] = int(cols[x+1])
            ability["value"] = int(cols[x+2])
            abilities.append(ability)
    result = json.dumps(config, indent=4)        
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()
    print "skill_star_config, end"


def main_quest_config():
    print "main_quest_config, start"
    outPath = getOutPath("mainQuest")
    ws = getWorkBook().get_sheet_by_name("MainQuest")
    config = collections.OrderedDict()
    startRowNum = 4
    i = 0
    tasks = collections.OrderedDict()
    sections = collections.OrderedDict()
    for row in ws.rows:
        i = i + 1
        cols = []
        if i < startRowNum:
            continue
        for cell in row:
            cols.append(cell.value)
        task = collections.OrderedDict()
        tasks[int(cols[0])] = task
        task["taskId"] = int(cols[0])
        task["title"] = str(cols[1])
        task["desc"] = unicode(cols[2])
        task["num"] = int(cols[3])
        task["normalRewards"] = json.loads(cols[4])
        task["chestRewards"] = json.loads(cols[5])
        task["type"] = int(cols[6])

        if cols[8]:
            section = collections.OrderedDict()
            sections[int(cols[8])] = section
            section["sectionId"] = int(cols[8])
            section["sortId"] = int(cols[9])
            section["taskIds"] = json.loads(cols[10])
            section["rewards"] = json.loads(cols[11])
            section["honorId"] = int(cols[12])
            section["display"] = int(cols[13])
    config["tasks"] = tasks
    config["sections"] = sections
    result = json.dumps(config, indent=4, ensure_ascii=False)        
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()
    print "main_quest_config, end"


def ulevel_config():
    print "ulevel_config, start"
    outPath = getOutPath("ulevel")
    ws = getWorkBook().get_sheet_by_name("Level")
    uLevelConfig = collections.OrderedDict()
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i+1
        cols = []
        if i < startRowNum:
            continue
        for cell in row:
            cols.append(cell.value)
        oneLevel = collections.OrderedDict()
        uLevelConfig[int(cols[0])] = oneLevel
        oneLevel["500"] = cols[1]
        oneLevel["2000"] = cols[2]
        oneLevel["5000"] = cols[3]
    result = json.dumps(uLevelConfig, indent=4)
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()
    print "ulevel_config, end"

    
def cmptt_task_config():
    print "cmptt_task_config, start"
    outPath = getOutPath("cmpttTask")
    ws = getWorkBook().get_sheet_by_name("Cmptt")
    config = collections.OrderedDict()
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i+1
        cols = []
        if i < startRowNum:
            continue
        for cell in row:
            cols.append(cell.value)
        one = collections.OrderedDict()
        if str(cols[0]) in config:
            raise KeyError("taskId %d repeat" % int(cols[0]))
        if cols[0] == None:
            break
        config[str(cols[0])] = one
        one["taskId"] = cols[0]
        one["fishPool"] = cols[1]
        assert int(cols[2]) in [1, 2]
        one["taskType"] = cols[2]
        one["desc"] = unicode(cols[3] or "")
        one["timeLong"] = cols[4]
        one["chestReward"] = cols[11].split("|")
        # one["coinReward"] = cols[10]
        # one["pearlReward"] = cols[11]
        # one["couponReward"] = cols[12]
        one["readySeconds"] = 10
        targets = collections.OrderedDict()
        one["targets"] = targets
        if cols[5]:
            targets["target1"] = cols[5]
        if cols[6]:
            targets["number1"] = cols[6]
        if cols[5] and cols[7]:
            targets["inter1"] = cols[7]
        if cols[8]:
            targets["target2"] = cols[8]
        if cols[9]:
            targets["number2"] = cols[9]
        if cols[8] and cols[10]:
            targets["inter2"] = cols[10]
    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()
    print "cmptt_task_config, end"


def ncmptt_task_config():
    print "ncmptt_task_config, start"
    outPath = getOutPath("ncmpttTask")
    ws = getWorkBook().get_sheet_by_name("Ncmptt")
    config = collections.OrderedDict()
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i+1
        cols = []
        if i < startRowNum:
            continue
        for cell in row:
            cols.append(cell.value)
        one = collections.OrderedDict()
        if str(cols[0]) in config:
            raise KeyError("taskId %d repeat" % int(cols[0]))
        config[str(cols[0])] = one
        one["taskId"] = cols[0]
        one["fishPool"] = cols[1]
        assert int(cols[2]) in [1, 2, 3, 4, 5]
        one["taskType"] = cols[2]
        one["desc"] = unicode(cols[3] or "")
        one["timeLong"] = cols[4]
        one["chestReward"] = cols[11]
        # one["coinReward"] = cols[10]
        one["readySeconds"] = 3
        targets = collections.OrderedDict()
        one["targets"] = targets
        if cols[5]:
            targets["target1"] = cols[5]
        if cols[6]:
            targets["number1"] = cols[6]
        if cols[5] and cols[7]:
            targets["inter1"] = cols[7]
        if cols[8]:
            targets["target2"] = cols[8]
        if cols[9]:
            targets["number2"] = cols[9]
        if cols[8] and cols[10]:
            targets["inter2"] = cols[10]
    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()
    print "ncmptt_task_config, end"


def bonus_task_config():
    print "bonus_task_config, start"
    outPath = getOutPath("bonusTask")
    ws = getWorkBook().get_sheet_by_name("Bonus")
    config = collections.OrderedDict()
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i+1
        cols = []
        if i < startRowNum:
            continue
        for cell in row:
            cols.append(cell.value)
        if len(cols) < 7 or cols[0] is None:
            continue
        one = collections.OrderedDict()
        if str(cols[0]) in config:
            raise KeyError("taskId %d repeat" % int(cols[0]))
        config[str(cols[0])] = one
        one["taskId"] = cols[0]
        one["fishPool"] = cols[1]
        assert int(cols[2]) in [1, 2, 4]
        one["taskType"] = cols[2]
        one["desc"] = unicode(cols[3] or "")
        one["timeLong"] = cols[4]
        # one["firstChestReward"] = cols[7]
        # one["secondChestReward"] = cols[8]
        one["readySeconds"] = 10
        targets = collections.OrderedDict()
        one["targets"] = targets
        if cols[5]:
            targets["target1"] = cols[5]
        if cols[5] and cols[6]:
            targets["inter1"] = cols[6]
        if cols[7]:
            targets["target2"] = cols[7]
        if cols[7] and cols[8]:
            targets["inter2"] = cols[8]
    result = json.dumps(config, indent=4, ensure_ascii=False)        
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close() 
    print "bonus_task_config, end"


def guide_task_config():
    print "guide_task_config, start"
    outPath = getOutPath("guideTask")
    ws = getWorkBook().get_sheet_by_name("Guide")
    config = collections.OrderedDict()
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i+1
        cols = []
        if i < startRowNum:
            continue
        for cell in row:
            cols.append(cell.value)
        one = collections.OrderedDict()
        if str(cols[0]) in config:
            raise KeyError("taskId %d repeat" % int(cols[0]))
        config[str(cols[0])] = one
        one["taskId"] = cols[0]
        assert int(cols[1]) in [1, 2, 3, 4, 5, 6]
        one["taskType"] = cols[1]
        one["coinReward"] = cols[7]
        one["pearlReward"] = cols[8]
        one["skillReward"] = cols[9]
        one["chestReward"] = cols[10]
        one["steps"] = eval(cols[11])
        targets = collections.OrderedDict()
        one["targets"] = targets
        if cols[3]:
            targets["target1"] = cols[3]
        if cols[4]:
            targets["number1"] = cols[4]
        if cols[5]:
            targets["target2"] = cols[5]
        if cols[6]:
            targets["number2"] = cols[6]
    result = json.dumps(config, indent=4)
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()
    print "guide_task_config, end"


def daily_quest_config():
    print "daily_quest_config, start"
    outPath = getOutPath("dailyQuest")
    ws = getWorkBook().get_sheet_by_name("DailyQuest")
    conf = collections.OrderedDict()
    conf["questData"] = collections.OrderedDict()
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i+1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if not cols[0]:
            continue
        one = collections.OrderedDict()
        conf["questData"][cols[0]] = one
        one["taskId"] = cols[0]
        one["taskType"] = cols[1]
        one["des"] = unicode(cols[2] or "")
        one["lv"] = cols[3]
        one["activeLv"] = cols[4]
        one["unlockUserLv"] = cols[5]
        one["taskLevel"] = cols[6]
        one["groupId"] = cols[7]
        one["targetsNum"] = cols[8]
        one["fishPool"] = json.loads(cols[9])
        one["fpMultiple"] = int(cols[10])
        showDay = cols[11]
        if isinstance(showDay, unicode):
            showDay = showDay.split("|")
            showDay = [int(x) for x in showDay]
        else:
            showDay = [showDay]
        one["showDay"] = showDay
        if cols[12]:
            one["rewards"] = json.loads(cols[12])

        if cols[14]:
            conf["questOrder"] = json.loads(cols[14])

    result = json.dumps(conf, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()
    daily_quest_reward_config()
    print "daily_quest_config, end"

def daily_quest_reward_config():
    outPath = getOutPath("dailyQuestReward")
    ws = getWorkBook().get_sheet_by_name("DailyQuestReward")
    conf = collections.OrderedDict()
    startRowNum = 4
    i = 0
    for row in ws.rows:
        i = i+1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if not cols[0]:
            continue
        conf[cols[0]] = collections.OrderedDict()
        for k in range(1, len(cols), 5):
            one = collections.OrderedDict()
            conf[cols[0]][cols[k]] = one
            one["finishedStar"] = cols[k]
            one["type"] = str(cols[k + 1])
            one["rewards"] = [{"itemId": cols[k + 3], "count": 1, "name": cols[k + 2], "reward": json.loads(cols[k + 4])}]

    result = json.dumps(conf, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
   

def getCouponFish(ws):
    couponFish = collections.OrderedDict()
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if not cols[0]: 
            continue
        one = collections.OrderedDict()
        couponFish[str(cols[0])] = one
        one["fishPool"] = int(cols[0])
        one["totalBullet"] = int(cols[1])
        one["minSecond"] = int(cols[2])
        one["maxSecond"] = int(cols[3])
        one["couponRange"] = json.loads(cols[4])
        fishes = []
        probb = 0
        for x in xrange(5, len(cols), 2):
            if not cols[x] or not cols[x + 1]:
                continue
            fish = {}
            fish["fishType"] = int(cols[x])
            itemProbb = int(cols[x + 1])
            fish["probb"] = [probb + 1, probb + itemProbb] 
            probb += itemProbb
            fishes.append(fish)
        one["fishes"] = fishes
    return couponFish


def getMultipleFish(ws):
    multipleFish = collections.OrderedDict()
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if not cols:
            continue
        if not cols[0]: 
            continue
        one = []
        multipleFish[str(cols[0])] = one
        probb = 0
        highProbb = 0
        rechargeProbb = 0
        for x in xrange(1, len(cols), 4):
            multiple = collections.OrderedDict()
            itemProbb = int(cols[x + 1])
            itemHighProbb = int(cols[x + 2])
            itemRechargeProbb = int(cols[x + 3])
            if itemProbb <= 0 and itemHighProbb <= 0 and itemRechargeProbb <= 0:
                continue
            multiple["multiple"] = int(cols[x])
            if itemProbb:
                multiple["probb"] = [probb + 1, probb + itemProbb]
            if itemHighProbb:
                multiple["highProbb"] = [highProbb + 1, highProbb + itemHighProbb]
            if itemRechargeProbb:
                multiple["rechargeProbb"] = [rechargeProbb + 1, rechargeProbb + itemRechargeProbb]
            probb += itemProbb
            highProbb += itemHighProbb
            rechargeProbb += itemRechargeProbb
            one.append(multiple)
    return multipleFish


# def getHitBoss(ws):
    # hitBoss = []
    # startRowNum = 4
    # i = 0
    # for row in ws.rows:
    #     i = i + 1
    #     if i < startRowNum:
    #         continue
    #     cols = []
    #     for cell in row:
    #         cols.append(cell.value)
    #     if not cols[0]: 
    #         continue
    #     probb = 0 
    #     for x in xrange(0, len(cols), 2):
    #         if not cols[x]:
    #             continue
    #         one = collections.OrderedDict()
    #         oneProbb = int(cols[x])
    #         one["probb"] = [probb + 1, probb + oneProbb]
    #         one["multiple"] = float(cols[x + 1])
    #         probb += oneProbb
    #         hitBoss.append(one)
    # return hitBoss


def getBossFish(ws):
    bossFish = collections.OrderedDict()
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if not cols[0]: 
            continue
        one = []
        bossFish[str(cols[0])] = one
        probb = 0
        for x in xrange(1, len(cols), 2):
            if not cols[x] or not cols[x + 1]:
                continue
            fish = {}
            fish["fishType"] = int(cols[x])
            itemProbb = int(cols[x + 1])
            fish["probb"] = [probb + 1, probb + itemProbb] 
            probb += itemProbb
            one.append(fish)
    return bossFish


def getChestFish(ws):
    chestFish = collections.OrderedDict()
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if not cols[0]: 
            continue
        one = collections.OrderedDict()
        chestFish[str(cols[0])] = one
        one["maxCoin"] = cols[1]
        one["minCoin"] = cols[2]
        one["probbs"] = []
        probb = 0
        for x in xrange(3, len(cols), 2):
            if not cols[x] or not cols[x + 1]:
                continue
            fish = {}
            fish["score"] = int(cols[x])
            itemProbb = int(cols[x + 1])
            fish["probb"] = [probb + 1, probb + itemProbb] 
            probb += itemProbb
            one["probbs"].append(fish)
    return chestFish


def getActivityFish(ws):
    activityFish = collections.OrderedDict()
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if not cols[0]: 
            continue
        one = collections.OrderedDict()
        activityFish[str(cols[0])] = one
        one["fishPool"] = int(cols[0])
        one["totalBullet"] = int(cols[1])
        one["minSecond"] = int(cols[2])
        one["maxSecond"] = int(cols[3])
        fishes = []
        probb = 0
        for x in xrange(4, len(cols), 2):
            if not cols[x] or not cols[x + 1]:
                continue
            fish = {}
            fish["fishType"] = int(cols[x])
            itemProbb = int(cols[x + 1])
            fish["probb"] = [probb + 1, probb + itemProbb] 
            probb += itemProbb
            fishes.append(fish)
        one["fishes"] = fishes
    return activityFish


def getShareFish(ws):
    shareFish = collections.OrderedDict()
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if not cols[0]: 
            continue
        one = collections.OrderedDict()
        shareFish[str(cols[0])] = one
        one["fishPool"] = int(cols[0])
        one["totalBullet"] = int(cols[1])
        one["minSecond"] = int(cols[2])
        one["maxSecond"] = int(cols[3])
        fishes = []
        probb = 0
        for x in xrange(4, len(cols), 2):
            if not cols[x] or not cols[x + 1]:
                continue
            fish = {}
            fish["fishType"] = int(cols[x])
            itemProbb = int(cols[x + 1])
            fish["probb"] = [probb + 1, probb + itemProbb] 
            probb += itemProbb
            fishes.append(fish)
        one["fishes"] = fishes
    return shareFish


def getAutofillFish(ws):
    autofillFish = collections.OrderedDict()
    startRowNum = 4
    i = 0
    isFindBlank = False
    for row in ws.rows:
        i = i + 1
        cols = []
        if not isFindBlank and i >= startRowNum:
            for cell in row:
                cols.append(cell.value)
            if not cols[0]:
                isFindBlank = True
                continue
            one = collections.OrderedDict()
            autofillFish[str(cols[0])] = []
            autofillFish[str(cols[0])].append(one)
            one["supplyInterval"] = int(cols[1])
            one["minCount"] = int(cols[2])
            one["maxCount"] = int(cols[3])
            one["minInterval"] = int(cols[4])
            one["maxInterval"] = int(cols[5])
            one["fish"] = []
            probb = 0
            for x in xrange(6, len(cols), 2):
                if not cols[x]:
                    continue
                fish = {}
                fish["fishType"] = int(cols[x])
                itemProbb = int(cols[x + 1])
                fish["probb"] = [probb + 1, probb + itemProbb]
                probb += itemProbb
                one["fish"].append(fish)

        if isFindBlank:
            for cell in row:
                cols.append(cell.value)
            if not cols[0]:
                continue
            try:
                supplyInterval = int(cols[1])
            except:
                continue
            for x in xrange(2, len(cols), 5):
                if not cols[x]:
                    continue
                one = collections.OrderedDict()
                fish = {}
                fish["fishType"] = int(cols[x])
                fish["probb"] = [0, 10000]
                one["fish"] = []
                one["fish"].append(fish)
                one["minCount"] = int(cols[x + 1])
                one["maxCount"] = int(cols[x + 2])
                one["minInterval"] = int(cols[x + 3])
                one["maxInterval"] = int(cols[x + 4])
                one["supplyInterval"] = supplyInterval
                autofillFish[str(cols[0])].append(one)
    return autofillFish


def getHippoFish(ws):
    hippoFish = collections.OrderedDict()
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if not cols[0]: 
            continue
        one = {}
        hippoFish[str(cols[0])] = one
        one["fishType"] = int(cols[1])
        one["minTime"] = int(cols[2])
        one["maxTime"] = int(cols[3])
        one["rewards"] = []
        probb = 0
        for x in xrange(4, len(cols), 3):
            if not cols[x] or not cols[x + 1] or not cols[x + 2]:
                continue
            itemInfo = {}
            itemInfo["name"] = int(cols[x])
            itemInfo["count"] = int(cols[x + 1])
            itemProbb = int(cols[x + 2])
            itemInfo["probb"] = [probb + 1, probb + itemProbb] 
            probb += itemProbb
            one["rewards"].append(itemInfo)
    return hippoFish


def getUserCouponFish(ws):
    userCouponFish = collections.OrderedDict()
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if not cols[0]: 
            continue
        one = collections.OrderedDict()
        userCouponFish.setdefault(str(cols[0]), []).append(one)
        # userCouponFish[str(cols[0])] = one
        one["fishPool"] = int(cols[0])
        one["limitCount"] = int(cols[1])
        one["sections"] = json.loads(cols[2])
        one["totalBullet"] = int(cols[3])
        one["minSecond"] = int(cols[4])
        one["maxSecond"] = int(cols[5])
        fishes = []
        probb = 0
        for x in xrange(6, len(cols), 2):
            if not cols[x] or not cols[x + 1]:
                continue
            fish = {}
            fish["fishType"] = int(cols[x])
            itemProbb = int(cols[x + 1])
            fish["probb"] = [probb + 1, probb + itemProbb] 
            probb += itemProbb
            fishes.append(fish)
        one["fishes"] = fishes
    return userCouponFish

def probability_config():
    outPath = getOutPath("probability")
    wb = getWorkBook()
    config = collections.OrderedDict()
    couponFish = getCouponFish(wb.get_sheet_by_name("CouponFish"))
    multipleFish = getMultipleFish(wb.get_sheet_by_name("MultipleFish"))
    # hitBoss = getHitBoss(wb.get_sheet_by_name("HitBoss"))
    bossFish = getBossFish(wb.get_sheet_by_name("BossFish"))
    chestFish = getChestFish(wb.get_sheet_by_name("ChestFish"))
    activityFish = getActivityFish(wb.get_sheet_by_name("ActivityFish"))
    shareFish = getShareFish(wb.get_sheet_by_name("ShareFish"))
    autofillFish = getAutofillFish(wb.get_sheet_by_name("AutofillFish"))
    hippoFish = getHippoFish(wb.get_sheet_by_name("HippoFish"))
    userCouponFish = getUserCouponFish(wb.get_sheet_by_name("UserCouponFish"))

    config["couponFish"] = couponFish
    config["multipleFish"] = multipleFish
    config["bossFish"] = bossFish
    # config["hitBoss"] = hitBoss
    config["chestFish"] = chestFish
    config["activityFish"] = activityFish
    config["shareFish"] = shareFish
    config["autofillFish"] = autofillFish
    config["hippoFish"] = hippoFish
    config["userCouponFish"] = userCouponFish
    result = json.dumps(config, indent=4)        
    outHandle = open(outPath, "w")   
    outHandle.write(result)  


def terror_fish():
    """特殊鱼出现的概率和间隔"""
    outPath = getOutPath("terrorFish")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("TerrorFish")
    config = collections.OrderedDict()
    terrorFish = collections.OrderedDict()
    startRowNum = 4
    i = 0
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if not cols[0]:
            continue
        one = []
        terrorFish[str(cols[0])] = one
        interval = int(cols[1])
        probb = 0
        for x in xrange(2, len(cols), 2):
            if not cols[x] or not cols[x + 1]:
                continue
            fish = {}
            fish["fishType"] = int(cols[x])
            fish["interval"] = interval
            itemProbb = int(cols[x + 1])
            fish["probb"] = [probb + 1, probb + itemProbb]
            probb += itemProbb
            one.append(fish)
    config["terrorFish"] = terrorFish
    result = json.dumps(config, indent=4)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()


def dynamic_odds_config():
    print u"start 动态概率配置 DynamicOdds"
    outPath = getOutPath("dynamicOdds")
    ws = getWorkBook().get_sheet_by_name("DynamicOdds")
    dynamicOdds = collections.OrderedDict()
    startRowNum = 4
    i = 0
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if not cols[0]: 
            continue
        dynamicOdds.setdefault(int(cols[0]), collections.OrderedDict())     # 所在渔场
        fishPool = dynamicOdds[int(cols[0])]
        wave = collections.OrderedDict()
        fishPool[str(cols[1])] = wave   # 曲线ID
        wave["waveId"] = int(cols[1])   # 曲线ID
        wave["weight"] = int(cols[2])   # 权重
        wave["type"] = float(cols[3])   # 涨还是跌 0为分界线
        wave["frequency"] = []          # 曲率列表
        for x in xrange(4, len(cols)):
            if cols[x] is None:
                break
            wave["frequency"].append(float(cols[x]))

    result = json.dumps(dynamicOdds, indent=4)        
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()
    print u"end 动态概率配置 DynamicOdds"


def lottery_pool_config():
    outPath = getOutPath("lotteryPool")
    ws = getWorkBook().get_sheet_by_name("LotteryPool")
    lotteryPool = collections.OrderedDict()
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if not cols[0]: 
            continue
        one = collections.OrderedDict()
        lotteryPool[cols[0]] = one
        one["lotteryPool"] = int(cols[1])
        one["skillPool"] = int(cols[2])
        one["multiplePool"] = int(cols[3])
        one["chestPool"] = int(cols[4])
        one["rainbowPool"] = int(cols[5])
        one["bonusRatio"] = cols[6]
        one["skillRatio"] = cols[7]
        one["multipleRatio"] = cols[8]
        one["bossRatio"] = cols[9]
        one["cmpttRatio"] = cols[10]
        one["chestRatio"] = cols[11]
        one["gameTimeRatio"] = cols[12]
        one["rainbowRatio"] = cols[13]
        one["grandPrizeRatio"] = cols[14]
        one["alertThreshold"] = cols[15]
    result = json.dumps(lotteryPool, indent=4)        
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()


def activity_config():
    reload(sys)
    sys.setdefaultencoding("utf-8")
    outPath = getOutPath("activity")
    ws = getWorkBook().get_sheet_by_name("Activity")
    taskConfig = activity_task()
    config = collections.OrderedDict()
    config["activityInfo"] = collections.OrderedDict()
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i+1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if not cols[0]: 
            continue
        one = collections.OrderedDict()
        if str(cols[0]) in config:
            raise KeyError("activityId %d repeat" % int(cols[0]))
        one["Id"] = cols[0]
        one["tip"] = cols[1]
        one["name"] = cols[2]
        if cols[3]:
            one["rule"] = json.loads(cols[3])
        else:
            one["rule"] = {}
        one["type"] = int(cols[4])
        one["tabType"] = int(cols[5])
        one["tabName"] = str(cols[6])
        one["clientModel"] = int(cols[7])
        one["modelUIType"] = int(cols[8])
        one["isDailyReset"] = int(cols[9])
        one["limitVip"] = int(cols[10])
        one["limitLevel"] = int(cols[11])
        one["limitGuide"] = int(cols[12])
        one["limitGameTime"] = int(cols[13])
        one["registerTime"] = int(cols[14])
        one["reward"] = cols[15]
        one["order"] = int(cols[16])
        one["lowClientVersion"] = (cols[17])
        one["reviewVerLimit"] = int(cols[18])
        if cols[19] and str(cols[20]) != "0":
            one["sysLimit"] = json.loads(cols[19])
        if str(cols[20]) == "0":
            continue
        try:
            one["activityEnable"] = json.loads(cols[20])
        except ValueError:
            raise KeyError("activity_config %s" % cols[0])
        one["effectiveTime"] = 0
        if cols[21] and cols[21] != "0":
            one["effectiveTime"] = {}
            one["effectiveTime"]["start"] = cols[21]
            one["effectiveTime"]["end"] = cols[22]
        if cols[23] and cols[23] != "0":
            timeArr = str(cols[23]).split("|")
            one["realAcTime"] = {}
            one["realAcTime"]["start"] = timeArr[0]
            one["realAcTime"]["end"] = timeArr[1]
        one["overDay"] = int(cols[24])
        one["receivedVisible"] = int(cols[25])
        if cols[26] and cols[26] != "0":
            one["dayTimeLimit"] = int(cols[26])
        if cols[27] and cols[27] != "0":
            one["notInTable"] = int(cols[27])
        if cols[28] and cols[28] != "0":
            one["sort"] = int(cols[28])
        one["colors"] = cols[29]
        one["activityTag"] = int(cols[30])
        one["isShowModuleTip"] = int(cols[31])
        if cols[32]:
            try:
                one["extends"] = json.loads(cols[32])
            except:
                one["extends"] = str(cols[32])
        one["acImg"] = cols[33]
        if cols[34]:
            one["acImgApp"] = cols[34] or ""
        if cols[35]:
            one["acImgQQ"] = cols[35] or ""
        if cols[36]:
            one["acImgEN"] = cols[36] or ""
        one["buttonParams"] = []
        for x in xrange(37, len(cols), 5):
            if cols[x]:
                item = {}
                item["action"] = cols[x]
                item["btnStr"] = cols[x + 1]
                if cols[x + 2]:
                    item["btnStrEN"] = cols[x + 2]
                item["param"] = cols[x + 3]
                item["btnPos"] = cols[x + 4]
                one["buttonParams"].append(item)
            else:
                break
        taskKey = one["Id"][0:-9]
        if taskConfig.has_key(taskKey):
            one["task"] = taskConfig[taskKey]
        config["activityInfo"][str(cols[0])] = one
    config["notice"] = notice_config()
    config["activityClient"] = activity_Client()
    result = json.dumps(config, indent=4, ensure_ascii=False)        
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close() 

def notice_config():
    ws = getWorkBook().get_sheet_by_name("notice")
    config = collections.OrderedDict()
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i+1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if not cols[0]: 
            continue
        one = collections.OrderedDict()
        if str(cols[0]) in config:
            raise KeyError("activityId %d repeat" % int(cols[0]))
        one["Id"] = cols[0]
        one["name"] = cols[1]
        one["type"] = int(cols[2])
        one["limitVip"] = int(cols[3])
        one["limitLevel"] = int(cols[4])
        one["limitGuide"] = int(cols[5])
        one["limitGameTime"] = int(cols[6])
        one["registerTime"] = int(cols[7])
        one["reward"] = cols[8]
        one["order"] = int(cols[9])
        one["lowClientVersion"] = (cols[10])
        one["reviewVerLimit"] = int(cols[11])
        if str(cols[12]) == "0":
            continue
        try:
            one["enable"] = json.loads(cols[12])
        except ValueError:
            raise KeyError("notice_config %s" % cols[0])
        one["effectiveTime"] = 0
        if cols[13] and cols[13] != "0":
            one["effectiveTime"] = {}
            one["effectiveTime"]["start"] = cols[13]
            one["effectiveTime"]["end"] = cols[14]
        one["noticeTag"] = int(cols[15])
        one["img"] = cols[16]
        if cols[17]:
            one["imgEN"] = cols[17]
        one["buttonParams"] = []
        for x in xrange(18, len(cols), 4):
            if cols[x]:
                item = {}
                item["action"] = cols[x]
                item["btnStr"] = cols[x + 1]
                item["param"] = cols[x + 2]
                item["btnPos"] = cols[x + 3]
                one["buttonParams"].append(item)
        config[str(cols[0])] = one
    return config

def activity_task():
    ws = getWorkBook().get_sheet_by_name("ActivityTask")
    config = collections.OrderedDict()
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i+1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if not cols[0]: 
            continue
        one = collections.OrderedDict()
        if config.has_key(str(cols[1])) and str(cols[0]) in config[str(cols[1])]:
            raise KeyError("activityId %d repeat" % int(cols[0]))
        one["id"] = cols[0]
        one["frontTasks"] = cols[2]
        one["showForever"] = cols[3]
        one["repeat"] = cols[4]
        one["type"] = cols[5]
        if cols[6]:
            one["isInside"] = cols[6]
        one["value"] = cols[7]
        one["reward"] = cols[8]
        if cols[9]:
            one["spareReward"] = cols[9]
        one["taskDesc"] = cols[10]
        one["taskImg"] = cols[12]
        one["taskDisableImg"] = cols[13]
        one["totalCount"] = cols[14]
        one["takeTimesPerDay"] = cols[15]
       
        if str(cols[1]) not in config:
            config[str(cols[1])] = collections.OrderedDict()
        config[str(cols[1])][str(cols[0])] = one
    return config    
    # result = json.dumps(config, indent=4, ensure_ascii=False)        
    # return result

def activity_Template(acConfigs):
    ws = getWorkBook().get_sheet_by_name("ActivityTemplate")
    config = collections.OrderedDict()

    for col in xrange(1, ws.max_column+1, 2):
        one = []
        if not ws.cell(row=1, column=col).value:
            continue
        config[str(ws.cell(row=1, column=col).value)] = one
        for row in xrange(1, ws.max_row+1):
            if row == 1:
                continue
            else:
                acId = ws.cell(row=row, column=col).value
                if acId and str(acId) in acConfigs:
                    one.append(str(acId))
                else:
                    if acId:
                        print acId
                        raise KeyError("activity_Template %s" % acId)
    return config

def activity_Client():
    ws = getWorkBook().get_sheet_by_name("ClientId")
    config = collections.OrderedDict()
    zhTempDict = collections.OrderedDict()
    enTempDict = collections.OrderedDict()
    config2 = []
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i+1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if not cols[0]: 
            continue
        one = []
        clientId = str(cols[0])
        clientIdNum = getClientIdNum(clientId)
        if cols[1]:
            zhTempDict[clientIdNum] = cols[1]
        if cols[2]:
            enTempDict[clientIdNum] = cols[2]
        if int(cols[3]) == 1:
            config2.append(clientIdNum)
    config["zh"] = zhTempDict
    config["en"] = enTempDict
    outPath = getOutPath("common")
    outHandle = open(outPath, "r")
    conf = json.load(outHandle)
    outHandle.close()
    conf["reviewLimitClientIds"] = config2
    def sortDict(_conf):
        if isinstance(_conf, dict):
            conf_sorted = sorted(_conf.iteritems())
            _conf = collections.OrderedDict()
            for k, v in conf_sorted:
                if isinstance(v, dict):
                    v = sortDict(v)
                elif isinstance(v, list):
                    __v = [sortDict(_v) for _v in v]
                    v = __v
                _conf[k] = v
        return _conf
    conf = sortDict(conf)
    # conf_sorted = sorted(conf.items(), key=lambda k: k[0])
    # conf = collections.OrderedDict()
    # for k, v in conf_sorted:
    #     conf[k] = v
    result = json.dumps(conf, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    return config

def catch_drop_config():
    outPath = getOutPath("catchDrop")
    ws = getWorkBook().get_sheet_by_name("CatchDrop")
    config = collections.OrderedDict()
    startRowNum = 4
    config["groupType"] = collections.OrderedDict()
    config["fishDrop"] = collections.OrderedDict()
    config["dropGroup"] = collections.OrderedDict()

    dropTypeStartIdx = 0
    dropDataStartIdx = 4
    dropGroupStartIdx = 17

    i = 0 
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if cols[dropTypeStartIdx] is not None:
            config["groupType"][cols[dropTypeStartIdx]] = json.loads(cols[1])

        if cols[dropDataStartIdx]:
            dropItemList = []
            config["fishDrop"].setdefault(str(cols[dropDataStartIdx]), collections.OrderedDict())
            config["fishDrop"][str(cols[dropDataStartIdx])][str(cols[dropDataStartIdx + 1])] = dropItemList
            for x in xrange(dropDataStartIdx + 2, dropGroupStartIdx - 1, 5):
                if cols[x] is not None:
                    itemDict = collections.OrderedDict()
                    itemDict["dropGroupId"] = int(cols[x])
                    itemDict["dropGroupPoss"] = int(cols[x + 1])
                    itemDict["min"] = int(cols[x + 2])
                    itemDict["max"] = int(cols[x + 3])
                    itemDict["probability"] = int(cols[x + 4])
                    dropItemList.append(itemDict)

        if cols[dropGroupStartIdx]:
            config["dropGroup"].setdefault(int(cols[dropGroupStartIdx]), [])
            config["dropGroup"][int(cols[dropGroupStartIdx])].extend([json.loads(cols[dropGroupStartIdx + 1]),
                                                                      json.loads(cols[dropGroupStartIdx + 2])])

    result = json.dumps(config, indent=4, ensure_ascii=False)        
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close() 


def vip_config():
    print "vip_config, start"
    outPath = getOutPath("vip")
    reload(sys)
    sys.setdefaultencoding("utf-8")
    ws = getWorkBook().get_sheet_by_name("Vip")
    config = collections.OrderedDict()
    startRowNum = 4

    expStartIdx = 11
    i = 0
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if not cols[0]:
            if i != startRowNum:
                break
        one = collections.OrderedDict()
        config[str(cols[0])] = one
        one["vipLv"] = int(cols[0])
        one["vipPresentCount:1137"] = int(cols[1])      # 赠送珍珠数
        one["vipPresentCount:1193"] = int(cols[2])      # 赠送银弹头数
        one["vipPresentCount:1194"] = int(cols[3])      # 赠送金弹头数
        one["vipPresentCount:1426"] = int(cols[4])      # 赠送代购券数
        one["vipPresentCount:1408"] = int(cols[5])      # 赠送冷却道具数量
        one["vipPresentCount:1429"] = int(cols[6])      # 赠送紫水晶数量
        one["vipPresentCount:1430"] = int(cols[7])      # 赠送黄水晶数量
        one["vipPresentCount:1431"] = int(cols[8])      # 赠送五彩水晶数量
        one["vipReceiveCount:1194"] = int(cols[9])      # 可接收金珠数量
        one["vipReceiveCount:1193"] = int(cols[10])     # 可接收银珠数量
        one["vipExp"] = int(cols[expStartIdx])          # VIP经验值
        one["freeRouletteTimes"] = int(cols[expStartIdx + 1])   # 免费海星许愿次数
        one["dropPearlTotalCount"] = int(cols[expStartIdx + 2]) # 掉落珍珠累计总量限制
        one["dropPearlRatioLimit"] = cols[expStartIdx + 3]      # 掉落总量超出后Level表中系数
        one["vipDesc"] = unicode(cols[expStartIdx + 4]) or ""   # VIP描述
        one["vipGift"] = json.loads(cols[expStartIdx + 5])      # VIP礼包
        one["originalPrice"] = int(cols[expStartIdx + 6])       # VIP礼包原价
        one["price"] = int(cols[expStartIdx + 7])               # VIP礼包现价
        one["pearlMultiple"] = float(cols[expStartIdx + 8])     # 任务珍珠加成倍率
        one["matchAddition"] = float(cols[expStartIdx + 9])     # 比赛分数加成
        one["setVipShow"] = int(cols[expStartIdx + 10])         # 设置vip展示
        one["almsRate"] = float(cols[expStartIdx + 11])         # 救济金倍率
        one["autoSupply:101"] = int(cols[expStartIdx + 12])     # 每日金币补足
        one["initLuckyValue:44102"] = int(cols[expStartIdx + 13])   # 比赛初始幸运值
        one["inviterReward"] = json.loads(cols[expStartIdx + 14])   # 给邀请人的奖励
        one["contFire"] = int(cols[expStartIdx + 15])               # 是否可以连发
        one["enableChat"] = int(cols[expStartIdx + 16])             # 是否可以聊天
        one["limitChatTip"] = unicode(cols[expStartIdx + 17] or "") # 限制聊天提示
        one["convert1137ToDRate"] = json.loads(cols[expStartIdx + 18])  # 珍珠换钻石比例
        one["convert1429ToDRate"] = json.loads(cols[expStartIdx + 19])  # 紫水晶换钻石比例
        one["convert1430ToDRate"] = json.loads(cols[expStartIdx + 20])  # 黄水晶换钻石比例
        one["convert1431ToDRate"] = json.loads(cols[expStartIdx + 21])  # 五彩水晶换钻石比例
        one["grandPrixFreeTimes"] = int(cols[expStartIdx + 22])         # 大奖赛免费次数
        one["grandPrixAddition"] = float(cols[expStartIdx + 23])        # 大奖赛分数加成
        one["checkinRechargeBonus"] = int(cols[expStartIdx + 24])       # 签到增加奖池额度

    result = json.dumps(config, indent=4, ensure_ascii=False)
    result = re.sub(r"\\\\n", r"\\n", result)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "vip_config, end"


def fixed_multiple_fish():
    """固定倍率鱼"""
    outPath = getOutPath("fixedMultipleFish")
    ws = getWorkBook().get_sheet_by_name("FixedMultipleFish")
    config = collections.OrderedDict()
    startRowNum = 4
    i = 0
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if not cols[0]:
            continue
        probbConf = {}
        one = []
        config[str(cols[0])] = probbConf
        probbConf["range"] = json.loads(str(cols[1]))
        probbConf["probb"] = cols[2]
        probbConf["multiples"] = one
        probb = 0
        for x in xrange(3, len(cols), 2):
            if not cols[x] or not cols[x + 1]:
                continue
            multiple = collections.OrderedDict()
            itemProbb = int(cols[x + 1])
            multiple["multiple"] = int(cols[x])
            multiple["probb"] = [probb + 1, probb + itemProbb]
            probb += itemProbb
            one.append(multiple)
    result = json.dumps(config, indent=4)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()


def call_multiple_fish():
    outPath = getOutPath("callMultipleFish")
    ws = getWorkBook().get_sheet_by_name("CallMultipleFish")
    ws2 = getWorkBook().get_sheet_by_name("MatchCallMultipleFish")
    config = collections.OrderedDict()
    normalConfig = collections.OrderedDict()
    matchConfig = collections.OrderedDict()
    validGroupConfig = collections.OrderedDict()
    startRowNum = 4
    i = 0
    firstConfEndRowNum = 0
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if not cols[0]:
            firstConfEndRowNum = i
            continue
        if firstConfEndRowNum > 0:
            if i > firstConfEndRowNum + 3:
                validGroupConfig[int(cols[0])] = json.loads(cols[1])
        else:
            one = []
            normalConfig[str(cols[0])] = one
            for x in xrange(1, len(cols), 2):
                if not cols[x] or not cols[x + 1]:
                    continue
                fishConf = collections.OrderedDict()
                fishConf["fishType"] = int(cols[x])
                fishConf["weight"] = int(cols[x + 1])
                one.append(fishConf)

    firstConfEndRowNum = 0
    startRowNum = 4
    i = 0
    for row in ws2.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if not cols[0]:
            firstConfEndRowNum = i
            continue
        if firstConfEndRowNum > 0:
            if i > firstConfEndRowNum + 3:
                validGroupConfig[int(cols[0])] = json.loads(cols[1])
        else:
            one = []
            matchConfig[str(cols[0])] = one
            for x in xrange(1, len(cols), 2):
                if not cols[x] or not cols[x + 1]:
                    continue
                fishConf = collections.OrderedDict()
                fishConf["fishType"] = int(cols[x])
                fishConf["weight"] = int(cols[x + 1])
                one.append(fishConf)
    config["normal"] = normalConfig
    config["match"] = matchConfig
    config["validGroup"] = validGroupConfig
    result = json.dumps(config, indent=4)        
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()


def fishBonus_config():
    outPath = getOutPath("fishBonusGame")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("StarfishBonus")
    wsGameTime = wb.get_sheet_by_name("GameTimeBonus")
    config = collections.OrderedDict()
    startRowNum = 4
    i = 0 
    config["starfishBonus"] = []
    config["gameTimeBonus"] = gameTime_bonus(wsGameTime)
    config["1283"] = eggs_bonus("egg1283")
    config["1284"] = eggs_bonus("egg1284")
    config["1285"] = eggs_bonus("egg1285")
    config["1286"] = eggs_bonus("egg1286")
    config["1287"] = eggs_bonus("egg1287")
    config["exchangeBonus"] = []
    config["eggs_dropConf"], config["eggs_minCoinRange"] = eggs_bonus_drop_conf()
    config["adBonus"] = ad_bonus()
    config["exchangeBonus"] = common_bonus("exchangeBonus")
    config["bulletPoolBonus"] = bullet_bet_bonus("BulletDouble")

    for row in ws.rows:
        i = i+1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if cols[0] !=0 and not cols[0]:
            continue
        one = collections.OrderedDict()
        config["starfishBonus"].append(one)
        one["Id"] = int(cols[0])
        one["weight"] = int(cols[1])
        one["clearLucky"] = int(cols[2])
        one["rewards"] = []
        for x in xrange(3, len(cols), 3):
            item = {}
            item["name"] = int(cols[x + 0])
            item["count"] = int(cols[x + 1])
            item["rare"] = int(cols[x + 2])
            one["rewards"].append(item)

    result = json.dumps(config, indent=4, ensure_ascii=False)        
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()

def bullet_bet_bonus(sheetName):
    ws = getWorkBook("eggsProbability.xlsx").get_sheet_by_name(sheetName)
    i = 0 
    results = collections.OrderedDict()
    startRowNum = 3
    for row in ws.rows:
        i = i+1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if cols[0] !=0 and not cols[0]:
            continue
        one = collections.OrderedDict()
        results[cols[0]] = one
        one["num"] = int(cols[0])
        one["probabilitys"] = []
        for x in xrange(1, len(cols), 2):
            item = {}
            item["bet"] = int(cols[x + 0])
            item["weight"] = float(cols[x + 1]) * 100
            one["probabilitys"].append(item)
    return results 

def gameTime_bonus(wsGameTime):
    rowNum = 4
    i = 0 
    result = []
    for row in wsGameTime.rows:
        i = i+1
        if i < rowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if cols[0] !=0 and not cols[0]:
            continue
        one = collections.OrderedDict()
        result.append(one)
        one["weight"] = int(cols[0])
        one["minCoin"] = int(cols[1])
        one["maxCoin"] = int(cols[2])
        one["luckyBoy"] = int(cols[3])
    return result

def common_bonus(sheetName):
    ws = getWorkBook("eggsProbability.xlsx").get_sheet_by_name(sheetName)
    i = 0 
    results = []
    startRowNum = 4
    for row in ws.rows:
        i = i+1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if cols[0] !=0 and not cols[0]:
            continue
        one = collections.OrderedDict()
        results.append(one)
        one["Id"] = int(cols[0])
        one["weight"] = int(cols[1])
        one["rewards"] = []
        for x in xrange(2, len(cols), 3):
            item = {}
            item["name"] = int(cols[x + 0])
            item["count"] = int(cols[x + 1])
            item["rare"] = int(cols[x + 2])
            one["rewards"].append(item)
    return results

def ad_bonus():
    ws = getWorkBook("eggsProbability.xlsx").get_sheet_by_name("adBonus")
    i = 0 
    results = []
    startRowNum = 4
    for row in ws.rows:
        i = i+1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if cols[0] !=0 and not cols[0]:
            continue
        one = collections.OrderedDict()
        results.append(one)
        one["Id"] = int(cols[0])
        one["weight"] = int(cols[1])
        one["weight1"] = int(cols[2])
        one["rewards"] = []
        for x in xrange(3, len(cols), 4):
            item = {}
            item["name"] = int(cols[x + 0])
            item["count"] = int(cols[x + 1])
            item["rare"] = int(cols[x + 2])
            one["rewards"].append(item)
    return results


def eggs_bonus(sheetName):
    ws = getWorkBook("eggsProbability.xlsx").get_sheet_by_name(sheetName)
    rowNum = 4
    i = 0 
    result = []
    for row in ws.rows:
        i = i+1
        if i < rowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if cols[0] !=0 and not cols[0]:
            continue
        one = collections.OrderedDict()
        result.append(one)
        one["minCoin"] = int(cols[0])
        one["maxCoin"] = int(cols[1])
        one["weight"] = int(cols[2])
    return result


def eggs_bonus_drop_conf():
    ws = getWorkBook("eggsProbability.xlsx").get_sheet_by_name("dropConf")
    rowNum = 4
    i = 0
    result = collections.OrderedDict()
    result2 = collections.OrderedDict()
    for row in ws.rows:
        i = i+1
        if i < rowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if cols[0] !=0 and not cols[0]:
            continue
        id = int(cols[0])
        result[id] = []
        for i in range(2, 26, 4):
            one = collections.OrderedDict()
            result[id].append(one)
            one["minCoin"] = int(cols[i + 1])
            one["maxCoin"] = int(cols[i + 2])
            one["count"] = int(cols[i + 3])
            one["color"] = cols[i + 4]
        result2[id] = [int(cols[1]), int(cols[2])]
    return result, result2


def getClientIdMap():
    import urllib2
    import urllib
    import json
    from hashlib import md5
    def dohttpquery(posturl, datadict):
        Headers = {"Content-type": "application/x-www-form-urlencoded"}
        postData = urllib.urlencode(datadict)
        request = urllib2.Request(url=posturl, data=postData, headers=Headers)
        response = urllib2.urlopen(request)
        if response != None :
            retstr = response.read()
            return retstr
        return "{}"
    ct = int(time.time())
    sign = "gdss.touch4.me-api-" + str(ct) + "-gdss.touch4.me-api"
    m = md5()
    m.update(sign)
    md5code = m.hexdigest()
    sign = md5code.lower()
    posturl = "%s/?act=api.%s&time=%d&sign=%s" % ("http://gdss.touch4.me", "getClientIdDict", ct, sign)
    retstr = dohttpquery(posturl, {})
    return json.loads(retstr).get("retmsg", {})


def getClientIdNum(clientId):
    global ClientIdMap
    ClientIdMap = ClientIdMap or getClientIdMap()
    result = ClientIdMap.get(clientId)
    print clientId, result
    assert result
    return result


def match_multiple_fish():
    outPath = getOutPath("matchMultipleFish")
    ws = getWorkBook().get_sheet_by_name("MatchMultipleFish")
    config = collections.OrderedDict()
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if not cols[0]: 
            continue
        probbConf = []
        config[str(cols[0])] = probbConf
        for x in xrange(4):
            multiples = []
            luckyValueConf = collections.OrderedDict()
            luckyValueConf["luckyValue"] = int(cols[13 * x + 1])
            luckyValueConf["multiples"] = multiples
            probb = 0
            for y in xrange(13 * x + 2, 13 * (x + 1), 2):
                if not cols[y] or not cols[y + 1]:
                    continue
                multiple = collections.OrderedDict()
                multiples.append(multiple)
                multiple["multiple"] = int(cols[y])
                itemProbb = int(cols[y + 1])
                multiple["probb"] = [probb + 1, probb + itemProbb]
                probb += itemProbb
            probbConf.append(luckyValueConf)
    result = json.dumps(config, indent=4)        
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()


def weaponPowerRate_config():
    outPath = getOutPath("weaponPowerRate")
    ws = getWorkBook().get_sheet_by_name("WeaponPowerRate")
    powerRateConfig = collections.OrderedDict()
    startRowNum = 4
    i = 0
    for row in ws.rows:
        i = i+1
        cols = []
        if i < startRowNum:
            continue
        for cell in row:
            cols.append(cell.value)
        oneRate =  []
        if str(cols[0]) in powerRateConfig:
            raise KeyError("weaponId %d repeat" % int(cols[0]))
        powerRateConfig[str(cols[0])] = oneRate
        probb = 0
        for m in range(2, len(cols), 2):
            if not cols[m]:
                break              
            item = {}
            item["value"] = json.loads(cols[m])
            itemProbb = cols[m + 1]
            item["probb"] = [probb + 1, probb + itemProbb]   
            oneRate.append(item)
            probb += itemProbb
    result = json.dumps(powerRateConfig, indent=4)
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()


def weapon_stage_count_config():
    """
    武器的爆炸次数
    """
    print "weapon_stage_count_config, start"
    outPath = getOutPath("weaponStageCount")
    ws = getWorkBook().get_sheet_by_name("WeaponStageCount")
    config = collections.OrderedDict()
    startRowNum = 4
    i = 0
    for row in ws.rows:
        i = i + 1
        cols = []
        if i < startRowNum:
            continue

        for cell in row:
            cols.append(cell.value)

        if not cols[0]:
            continue

        if int(cols[0]) not in config:
            config[int(cols[0])] = []
        for x in xrange(2, len(cols)):
            if cols[x] is None:
                continue
            config[int(cols[0])].append(int(cols[x]))
        if sum(config[int(cols[0])]) != 10000:
            raise KeyError("weapon_stage_count_config probb error")

    result = json.dumps(config, indent=4)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "weapon_stage_count_config, end"


def gunskin_config(clientId=0):
    """皮肤炮配置"""
    sn = "Gun" if clientId == 0 else "Gun" + "_" + str(clientId)
    sn2 = "GunSkin" if clientId == 0 else "GunSkin" + "_" + str(clientId)
    fn = "0.json" if clientId == 0 else str(clientId) + ".json"
    outPath = getOutPath("gun", fn)     # 输出到gun文件夹下 0.json 26312.json
    ws = getWorkBook().get_sheet_by_name(sn)
    wsSkin = getWorkBook().get_sheet_by_name(sn2)
    startRowNum = 4
    i = 0
    config = collections.OrderedDict()
    gunIds = []
    config["gunIds"] = gunIds                       # 皮肤炮IDS
    config["gun"] = collections.OrderedDict()       # 皮肤炮配置
    config["skin"] = collections.OrderedDict()      # 皮肤炮皮肤
    gunConf = collections.OrderedDict()
    for row in ws.rows:
        i = i+1
        cols = []
        if i < startRowNum:
            continue
        for cell in row:
            cols.append(cell.value)
        if not cols[2]:
            continue
        if row[1].value not in config["gun"].keys():
            gunConf = collections.OrderedDict()
            config["gun"][row[1].value] = gunConf
        one = collections.OrderedDict()
        gunConf[str(row[2].value)] = one
        if int(cols[1]) not in gunIds:
            gunIds.append(int(cols[1]))
        one["gunId"] = int(cols[1])
        one["gunLevel"] = int(cols[2])
        one["name"] = cols[0]
        one["multiple"] = int(cols[3])
        one["unlockType"] = int(cols[4])
        one["unlockValue"] = int(cols[5])
        one["equipType"] = int(cols[6])
        one["equipValue"] = int(cols[7])
        one["exp"] = int(cols[8])
        one["totalExp"] = int(cols[9])
        one["effectAddition"] = float(cols[10])
        one["effectType"] = int(cols[11])
        one["effectProbb"] = int(cols[12])
        one["unlockDesc"] = unicode(cols[13] or "")
        one["equipDesc"] = unicode(cols[14] or "")
        one["unitPrice"] = int(cols[15])
        one["skins"] = json.loads(str(cols[16]))
        aloofOdds = []
        one["aloofOdds"] = aloofOdds
        for m in range(17, len(cols), 2):
            if cols[m] is None:
                break
            oddsMap = {}
            oddsMap["odds"] = cols[m]
            oddsMap["probb"] = cols[m + 1]
            aloofOdds.append(oddsMap)

    i = 0
    for row in wsSkin.rows:
        i = i+1
        cols = []
        if i < startRowNum:
            continue
        one = collections.OrderedDict()
        for cell in row:
            cols.append(cell.value)
        config["skin"][cols[0]] = one
        one["skinId"] = cols[0]
        one["gunId"] = cols[1]
        one["kindId"] = cols[2]
        one["consumeCount"] = cols[3]
    result = json.dumps(config, indent=4, ensure_ascii=False)        
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()


def achievement_config_old():
    outPath = getOutPath("achievement")
    ws = getWorkBook().get_sheet_by_name("AchievementTask")
    config = collections.OrderedDict()
    startRowNum = 4
    i = 0 

    for row in ws.rows:
        i = i+1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if cols[0] != 0 and not cols[0]: 
            continue
        if cols[6] not in config:
            config[cols[6]] = []
        one = collections.OrderedDict()
        one["taskId"] = cols[0]
        one["desc"] = cols[1]
        one["target"] = json.loads(cols[2])
        one["type"] = cols[3]
        one["isMax"] = cols[4]
        one["repeat"] = cols[5]
        one["groupId"] = cols[6]
        one["honorId"] = cols[7]
        one["norReward"] = _getRewardInfo(cols[8])
        one["chestReward"] = _getRewardInfo(cols[9])
        config[cols[6]].append(one)
    result = json.dumps(config, indent=4, ensure_ascii=False)        
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()


def plyerBuffer_config():
    outPath = getOutPath("playerBuffer")
    ws = getWorkBook().get_sheet_by_name("PlayerBuffer")
    config = collections.OrderedDict()
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i+1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if cols[0] ==0 or not cols[0]: 
            continue
        one = collections.OrderedDict()
        config[str(cols[0])] = one
        one["bufferId"] = int(cols[0])
        one["weaponIds"] = json.loads(cols[1])
        one["weaponType"] = json.loads(cols[2])
        one["cdReduce"] = float(cols[3])
        one["coinAdd"] = float(cols[4])
        one["powerAdd"] = float(cols[5])
        one["duration"] = int(cols[6])
        one["delayTime"] = float(cols[7])
    result = json.dumps(config, indent=4, ensure_ascii=False)        
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()


def randomMultipleFish_config():
    outPath = getOutPath("randomMultipleFish")
    ws = getWorkBook().get_sheet_by_name("RandomMultipleFish")
    config = collections.OrderedDict()
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i+1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if not cols[0]: 
            continue
        multiples = []
        config[str(cols[0])] = multiples
        totalProbb = 0
        for x in range(1, len(cols), 2):
            if not cols[x + 1]:
                continue
            one = collections.OrderedDict()
            one["fishType"] = int(cols[x])
            one["probb"] = [totalProbb + 1, totalProbb + cols[x + 1]]
            totalProbb += cols[x + 1]
            multiples.append(one)
    result = json.dumps(config, indent=4, ensure_ascii=False)        
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()


def guideFollowedTask_config():
    outPath = getOutPath("GuideFollowedTask")
    ws = getWorkBook().get_sheet_by_name("GuideFollowedTask")
    config = []
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i+1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if cols[0] !=0 and not cols[0]: 
            continue
        one = collections.OrderedDict()
        one["taskId"] = int(cols[0])
        one["desc"] = cols[1]
        one["type"] = int(cols[2])
        one["repeat"] = int(cols[3])
        one["target"] = json.loads(cols[4])
        one["reward"] = _getRewardInfo(cols[5])
        config.append(one)
    result = json.dumps(config, indent=4, ensure_ascii=False)        
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()  


def _getRewardInfo(reward):
    """获取奖励信息"""
    reward = json.loads(reward)
    rewardInfo = []
    for x in xrange(0, len(reward), 2):
        reward_ = {"name": reward[x], "count": reward[x+1]}
        rewardInfo.append(reward_)
    return rewardInfo


def gunLevel():
    """火炮等级配置"""
    outPath = getOutPath("gunLevel")
    ws = getWorkBook().get_sheet_by_name("GunLevel")
    config = collections.OrderedDict()
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if cols[0] != 0 and not cols[0]: 
            continue
        one = collections.OrderedDict()
        config[str(cols[0])] = one
        one["gunLevel"] = int(cols[0])
        upgradeItems = collections.OrderedDict()
        if int(cols[1]):
            upgradeItems[cols[1]] = cols[2]
        for x in range(3, 6, 2):
            if int(cols[x]):
                upgradeItems[cols[x]] = cols[x + 1]
        if upgradeItems:
            one["upgradeItems"] = upgradeItems
        _tmpIdx = 7
        one["successRate"] = int(cols[_tmpIdx])
        if int(cols[_tmpIdx]) < 10000:
            protectItems = collections.OrderedDict()
            one["protectItems"] = protectItems
            protectItems[cols[_tmpIdx + 1]] = int(cols[_tmpIdx + 2])
        returnItems = []
        probb = 0
        for x in range(_tmpIdx + 3, _tmpIdx + 12, 3):
            if int(cols[x]):
                item = collections.OrderedDict()
                item["kindId"] = cols[x]
                item["count"] = cols[x + 1]
                item["probb"] = [probb + 1, probb + cols[x + 2]]
                returnItems.append(item)
                probb += cols[x + 2]
        if returnItems:
            one["returnItems"] = returnItems
        if cols[_tmpIdx + 12]:
            one["interval"] = float(cols[_tmpIdx + 12])
        if cols[_tmpIdx + 13] is not None:
            one["levelAddition"] = float(cols[_tmpIdx + 13])
        one["levelValue"] = int(cols[_tmpIdx + 14])
    result = json.dumps(config, indent=4, ensure_ascii=False)        
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()


def tableTask():
    outPath = getOutPath("tableTask")
    ws = getWorkBook().get_sheet_by_name("TableTask")
    wsRoomTask = getWorkBook().get_sheet_by_name("RoomTask")
    config = collections.OrderedDict()
    config["tasks"] = collections.OrderedDict()
    config["roomTask"] = collections.OrderedDict()
    startRowNum = 2
    i = 0 
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if cols[0] !=0 and not cols[0]: 
            continue
        one = collections.OrderedDict()
        config["tasks"][str(cols[0])] = one
        one["taskId"] = int(cols[0])
        one["type"] = int(cols[1])
        one["desc"] = cols[2]
        one["target"] = int(cols[3])
        one["targetNum"] = int(cols[4])
        one["timeLong"] = int(cols[5])
        one["failTime"] = int(cols[6])
        one["rewards"] = []
        for m in range(7, 9, 3):
            if not cols[m]:
                break
            reward = {"name": int(cols[m]), "count": int(cols[m+1])}
            one["rewards"].append(reward)
        one["suggestTarget"] = int(cols[10])
        one["waitTime"] = int(cols[11])
    i = 0
    for row in wsRoomTask.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if cols[0] !=0 and not cols[0]: 
            continue
        if not cols[1] or cols[1] == "":
            continue
        one = collections.OrderedDict()
        config["roomTask"][str(cols[0])] = one
        one["mainTask"] = json.loads(cols[1])
        one["branchTask"] = json.loads(cols[2])
        one["randomTask"] = json.loads(cols[3])
    result = json.dumps(config, indent=4, ensure_ascii=False)        
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()  


def inviteTask_config():
    outPath = getOutPath("inviteTask")
    ws = getWorkBook().get_sheet_by_name("InviteTask")
    config = collections.OrderedDict()
    config["inviteTask"] = collections.OrderedDict()
    config["recallTask"] = collections.OrderedDict()
    startRowNum = 3
    i = 0 
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if cols[0] !=0 and not cols[0]: 
            continue
        one = collections.OrderedDict()
        if int(cols[1]) == 1:
            config["inviteTask"][str(cols[0])] = one
        else:
            config["recallTask"][str(cols[0])] = one
         
        one["Id"] = int(cols[0])
        one["type"] = int(cols[1])
        one["target"] = int(cols[2])
        one["rewards"] = []
        for m in range(3, len(cols), 2):
            if not cols[m]:
                break
            reward = {"name": int(cols[m]), "count": int(cols[m+1])}
            one["rewards"].append(reward)
    result = json.dumps(config, indent=4, ensure_ascii=False)        
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()


def rechargePool_config():
    outPath = getOutPath("rechargePool")
    ws = getWorkBook().get_sheet_by_name("RechargePool")
    config = collections.OrderedDict()
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if not cols[0]: 
            continue
        one = collections.OrderedDict()
        config[str(cols[0])] = one
        one["productId"] = str(cols[0])
        bonuses = []
        probb = 0
        for x in range(1, len(cols), 2):
            if not cols[x] or not cols[x + 1]:
                break
            bonus = collections.OrderedDict()
            bonus["bonus"] = int(cols[x])
            bonus["probb"] = [probb + 1, probb + int(cols[x + 1])]
            probb += int(cols[x + 1])
            bonuses.append(bonus)
        one["bonuses"] = bonuses
    result = json.dumps(config, indent=4, ensure_ascii=False)        
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()


def share_config():
    outPath = getOutPath("share")
    ws = getWorkBook().get_sheet_by_name("Share")
    config = collections.OrderedDict()
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if cols[0] !=0 and not cols[0]: 
            continue
        one = collections.OrderedDict()
        config[str(cols[0])] = one
        one["shareId"] = int(cols[0])
        one["typeId"] = str(cols[1])
        one["title"] = unicode(cols[2] or "")
        one["desc1"] = unicode(cols[3] or "")
        one["desc2"] = unicode(cols[4] or "")
        one["vipLimit"] = int(cols[6])
        one["levelLimit"] = int(cols[7])
        one["versionLimit"] = json.loads(cols[8])
        one["timeLimit"] = json.loads(cols[9])
        one["locationLimit"] = [unicode(x) for x in cols[10].split(",")] if cols[10] else []
        one["popCountLimit"] = int(cols[11])
        one["finishCountLimit"] = int(cols[12])
        one["expiresType"] = int(cols[13])
        one["finishType"] = int(cols[14])
        one["groupType"] = int(cols[15])
        one["modes"] = json.loads(cols[16])
        one["rewards1"] = json.loads(cols[17])
        one["rewards2"] = json.loads(cols[18])
    result = json.dumps(config, indent=4, ensure_ascii=False)        
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()


def fly_pig_config():
    outPath = getOutPath("flyPigReward")
    ws = getWorkBook().get_sheet_by_name("FlyPigReward")
    config = collections.OrderedDict()
    startRowNum = 4
    i = 0 
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if not cols[0]:
            break
        probb = 0
        items = []
        config[str(cols[0])] = items
        for x in range(1, len(cols), 6):
            if not cols[x] or not cols[x + 1]:
                continue
            itemDict = collections.OrderedDict()
            itemDict["probb"] = [probb + 1, probb + int(cols[x])]
            itemDict["kindId"] = int(cols[x + 1])
            itemDict["displayCount"] = int(cols[x + 2])
            itemDict["shareMultiple"] = int(cols[x + 3])
            itemDict["advertMultiple"] = int(cols[x + 4])
            itemDict["saleCoin"] = int(cols[x + 5])
            items.append(itemDict)
            probb += int(cols[x])
    result = json.dumps(config, indent=4, ensure_ascii=False)        
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()


def starfish_roulette_config():
    """
    海星轮盘配置
    """
    print "starfish_roulette_config, start"
    outPath = getOutPath("starfishRoulette")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("StarfishTurntable")
    config = collections.OrderedDict()

    config["starfishRoulette"] = []
    config["priceInfo"] = {}
    config["rouletteLevel"] = []
    config["unlockLevel"] = 0
    config["rule"] = ""

    # 轮盘配置数据范围
    # rouletteStartColIdx = 1
    rouletteEndColIdx = 7

    # 价格配置数据范围
    priceStartColIdx = 8
    priceEndColIdx = 13

    # 轮盘等级配置数据范围
    levelStartColIdx = 15
    levelEndColIdx = 17

    # 轮盘解锁等级
    unlockLevelColIdx = 19

    ruleIdx = 21

    startRowNum = 4
    i = 0
    lastLevelInfo = {"level": 0, "exp": 0, "count": 0}
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        if cols[0]:
            one = collections.OrderedDict()
            config["starfishRoulette"].append(one)
            one["Id"] = int(cols[0])
            one["starweight"] = int(cols[1])
            one["ticketweight"] = int(cols[2])
            one["rewards"] = []
            for x in xrange(3, rouletteEndColIdx, 4):
                item = {}
                item["name"] = int(cols[x + 0])
                item["count"] = int(cols[x + 1])
                item["rare"] = int(cols[x + 2])
                item["des"] = str(cols[x + 3])
                one["rewards"].append(item)

        if cols[priceStartColIdx]:
            priceCount = int(cols[priceStartColIdx])
            itemId = int(cols[priceStartColIdx + 1])
            if itemId:
                priceIdx = 0
                itemDict = {}
                for idx in xrange(priceStartColIdx + 2, priceEndColIdx, 2):
                    if priceIdx >= priceCount:
                        break
                    priceIdx = priceIdx + 1
                    item = {}
                    playTimes = int(cols[idx + 0])
                    item["name"] = int(itemId)
                    item["count"] = int(cols[idx + 1])
                    itemDict[playTimes] = item
                config["priceInfo"][itemId] = itemDict

        if cols[levelStartColIdx]:
            level = int(cols[levelStartColIdx])
            exp = int(cols[levelStartColIdx + 1])
            count = int(cols[levelStartColIdx + 2])
            if len(config["rouletteLevel"]) == 0:
                config["rouletteLevel"].append({"level": 0, "exp": 0, "count": 0})

            lastLevel = lastLevelInfo["level"]
            if level - lastLevel > 1:
                for idx in range(level - lastLevel - 1):
                    item = copy.deepcopy(lastLevelInfo)
                    item["level"] = lastLevel + idx + 1
                    config["rouletteLevel"].append(item)

            lastLevelInfo["level"] = level
            lastLevelInfo["exp"] = exp
            lastLevelInfo["count"] = count
            levelItem = copy.deepcopy(lastLevelInfo)
            config["rouletteLevel"].append(levelItem)

        if cols[unlockLevelColIdx]:
            config["unlockLevel"] = int(cols[unlockLevelColIdx])

        if cols[ruleIdx]:
            config["rule"] = str(cols[ruleIdx])

    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "starfish_roulette_config, end"


def surpass_target_config():
    """
    比赛超越目标配置
    """
    print "surpass_target_config, start"
    outPath = getOutPath("surpassTarget")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("SurpassTarget")
    config = []

    startRowNum = 4
    i = 0
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        if cols[0]:
            rank = int(cols[0])
            config.append(rank)

    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "surpass_target_config, end"


def special_item_config():
    """
    特殊物品配置
    """
    print "special_item_config, start"
    outPath = getOutPath("specialItem")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("SpecialItem")
    config = collections.OrderedDict()

    startRowNum = 4
    i = 0
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        if cols[0]:
            one = collections.OrderedDict()
            config[int(cols[0])] = one
            one["itemId"] = int(cols[0])
            one["incrPearlDropRate"] = cols[1]
            one["incrCrystalDropRate"] = cols[2]

    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "special_item_config, end"


def report_fish_cb_config():
    """
    上报捕鱼成本收益配置
    """
    print "report_fish_cb_config, start"
    outPath = getOutPath("reportFishCB")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("ReportFishCB")
    config = collections.OrderedDict()
    
    startRowNum = 4
    i = 0
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        
        if cols[0]:
            one = collections.OrderedDict()
            config[int(cols[0])] = one
            one["enable"] = int(cols[1])
            one["interval"] = cols[2]
            one["count"] = cols[3]

    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "report_fish_cb_config, end"


def robot_config():
    """
    机器人配置
    """
    print "robot_config, start"
    outPath = getOutPath("robot")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("Robot")
    config = collections.OrderedDict()
    outHandle = open(outPath, "r")
    ret = outHandle.read()
    outHandle.close()
    conf = json.loads(ret)
    conf["robotScript"] = config
    startRowNum = 4
    i = 0
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        if cols[0]:
            one = collections.OrderedDict()
            config[int(cols[0])] = one
            one["enable"] = int(cols[1])
            one["scripts"] = []
            for x in range(2, len(cols), 6):
                if cols[x] is None:
                    continue
                item = collections.OrderedDict()
                item["level"] = int(cols[x])
                item["fireRange"] = json.loads(cols[x + 1])
                item["idleRange"] = json.loads(cols[x + 2])
                item["leaveRange"] = json.loads(cols[x + 3])
                item["chatExp"] = json.loads(cols[x + 4])
                item["probb"] = int(cols[x + 5])
                one["scripts"].append(item)

    result = json.dumps(conf, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "robot_config, end"


def multi_lang_text():
    """
    多语言文本配置
    """
    print "multi_lang_text, start"
    reload(sys)
    sys.setdefaultencoding("utf-8")
    ws = getWorkBook("multiLangText.xlsx").get_sheet_by_name('multiLangText')
    config = collections.OrderedDict()
    outPath = getOutPath("multiLangText")
    i = 0 
    startRowNum = 4
    for row in ws.rows:
        i = i+1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if str(cols[0]) in config:
            raise KeyError("key %s repeat" % str(cols[0]))
        if cols[0] is not None:
            one = collections.OrderedDict()
            config[str(cols[0])] = one
            one["zh"] = str(cols[1])
            if cols[2]:
                one["en"] = str(cols[2])

    result = json.dumps(config, indent=4, ensure_ascii=False)
    result = re.sub(r"\\\\n", r"\\n", result)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "multi_lang_text, end"


def world_boss_config():
    """
    world boss配置
    """
    print "world_boss_config, start"
    outPath = getOutPath("worldboss")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("Worldboss")
    config = collections.OrderedDict()
    info = collections.OrderedDict()
    times = collections.OrderedDict()
    hp = []
    rewards = []
    hit_rewards = collections.OrderedDict()
    startRowNum_times = 4
    startRowNum_hp = 9
    startRowNum_rewards = 24
    i = 0
    for row in ws.rows:
        i = i + 1
        if startRowNum_hp - 4 > i >= startRowNum_times:
            cols = []
            for cell in row:
                cols.append(cell.value)

            if cols[0]:
                times["enable"] = int(cols[0])
                times["days"] = json.loads(cols[1])
                times["times_in_day"] = json.loads(cols[2])
                info["ready_interval"] = int(cols[3])
                info["room_list"] = json.loads(cols[4])
                info["fish_type"] = int(cols[5])
                info["life_time"] = int(cols[6])

        if startRowNum_rewards - 4 > i >= startRowNum_hp:
            cols = []
            for cell in row:
                cols.append(cell.value)

            if cols[0]:
                one = collections.OrderedDict()
                one["idx"] = int(cols[0])
                one["hp"] = int(cols[1])
                one["decrease_ratio"] = float(cols[2])
                one["lock_rate"] = float(cols[3])
                one["last_shot_rewards"] = json.loads(cols[5])
                one["vip_level"] = int(cols[6])
                one["hit_probb"] = int(cols[7])
                one["hit_rate"] = int(cols[8])
                one["drop"] = []
                for x in range(9, len(cols), 5):
                    one["drop"].append({"drop_probb": int(cols[x]),
                                        "item_id": int(cols[x + 2]),
                                        "item_count_min": int(cols[x + 3]),
                                        "item_count_max": int(cols[x + 4])})
                hp.append(one)

        if i >= startRowNum_rewards:
            cols = []
            for cell in row:
                cols.append(cell.value)

            if cols[0]:
                for x in range(2, len(cols), 2):
                    if cols[x] is None:
                        continue
                    rewards.append(json.loads(cols[x]))

    config["times"] = times
    config["hp"] = hp
    config["rank_rewards"] = rewards
    config["info"] = info
    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "world_boss_config, end"


def grand_prize_pool_config():
    """
    巨奖奖池配置
    """
    print "grand_prize_pool_config, start"
    outPath = getOutPath("grandPrize")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("GrandPrize")
    config = collections.OrderedDict()
    startRowNum = 4
    i = 0
    for row in ws.rows:
        i = i + 1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if cols[0]:
            config["roomIds"] = json.loads(cols[0])
            config["enable"] = int(cols[1])
            config["recordCount"] = int(cols[2])
            config["interval"] = json.loads(cols[3])
            config["winningBaseCoin"] = int(cols[4])
            config["minPool"] = int(cols[5])
            config["fishScore"] = int(cols[6])
            config["rewardsPct"] = []
            config["rewardsPct"].append(json.loads(cols[7]))
            config["rewardsPct"].append(json.loads(cols[8]))
            config["rewardsPct"].append(json.loads(cols[9]))
            if cols[10]:
                config["minWinningRate"] = int(cols[10])
            config["poolPct"] = cols[11]
            break
    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "grand_prize_pool_config, end"


def piggy_bank_config():
    """
    存钱罐池配置
    """
    reload(sys)
    sys.setdefaultencoding("utf-8")
    print "piggy_bank_config, start"
    outPath = getOutPath("piggyBank")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("PiggyBank")
    config = collections.OrderedDict()
    rule = {}
    startRowNum = 4
    h = 0
    for row in ws.rows:
        h = h + 1
        if h < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        for k in range(0, 4, 2):
            if cols[k]:
                rule[str(cols[k])] = unicode(cols[k + 1] or "")
        if cols[5] is None:
            break
        one = collections.OrderedDict()
        config[int(cols[5])] = one
        for i in range(6, len(cols), 17):
            one[str(cols[i])] = collections.OrderedDict()
            one[str(cols[i])]["type"] = str(cols[i])
            one[str(cols[i])]["productId"] = str(cols[i + 1])
            one[str(cols[i])]["productName"] = unicode("存钱罐")
            one[str(cols[i])]["initVal"] = int(cols[i + 2])
            if cols[i + 3] > 0:
                one[str(cols[i])]["maxDailyCount"] = int(cols[i + 3] * 10000)
            else:
                one[str(cols[i])]["maxDailyCount"] = int(cols[i + 3])
            if cols[i + 4] > 0:
                one[str(cols[i])]["maxCount"] = int(cols[i + 4] * 10000)
            else:
                one[str(cols[i])]["maxCount"] = int(cols[i + 4])
            one[str(cols[i])]["iscooling"] = int(cols[i + 5])
            one[str(cols[i])]["endcoolingTime"] = int(cols[i + 6])
            one[str(cols[i])]["price_direct"] = int(cols[i + 7])
            one[str(cols[i])]["price_diamond"] = int(cols[i + 8])
            one[str(cols[i])]["inroom"] = int(cols[i + 9])
            one[str(cols[i])]["outroom"] = int(cols[i + 10])
            one[str(cols[i])]["firePct"] = cols[i + 11]
            one[str(cols[i])]["buyType"] = unicode(cols[i + 12] or "")
            one[str(cols[i])]["otherBuyType"] = json.loads(cols[i + 13])
            one[str(cols[i])]["price"] = int(cols[i + 14])
            one[str(cols[i])]["resetTime"] = int(cols[i + 15])
            one[str(cols[i])]["rule"] = rule[str(cols[i])]
            one[str(cols[i])]["profitPct"] = cols[i + 16]
    result = json.dumps(config, indent=4, ensure_ascii=False)
    result = re.sub(r"\\\\n", r"\\n", result)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "piggy_bank_config, end"


def treasure_config():
    """
    宝藏配置
    """
    outPath = getOutPath("treasure")
    ws = getWorkBook().get_sheet_by_name("Treasure")
    config = collections.OrderedDict()
    startRowNum = 4
    i = 0
    one = collections.OrderedDict()
    levels = None
    for row in ws.rows:
        i = i+1
        cols = []
        if i < startRowNum:
            continue
        if row[1].value not in config.keys():
            one = collections.OrderedDict()
            config[row[1].value] = one
            levels = []
        for cell in row:
            cols.append(cell.value)
        if not cols[2]:
            continue
        one["name"] = unicode(cols[0])
        one["kindId"] = int(cols[1])
        one["sortId"] = int(cols[2])
        one["rare"] = int(cols[3])      
        one["effectType"] = int(cols[4])
        one["actions"] = json.loads(cols[5])
        one["limitCount"] = int(cols[6])
        one["desc"] = str(cols[7])
        one["effectDesc"] = str(cols[8])
        levelMap = collections.OrderedDict()
        levelMap["level"] = int(cols[9])
        levelMap["holdNum"] = int(cols[10])
        levelMap["params"] = json.loads(cols[11])
        levels.append(levelMap)
        one["levels"] = levels
    result = json.dumps(config, indent=4, ensure_ascii=False)        
    outHandle = open(outPath, "w")   
    outHandle.write(result)  
    outHandle.close()


def time_limited_store_config():
    """
    限时商城配置
    """
    reload(sys)
    sys.setdefaultencoding("utf-8")
    print "time_limited_store_config, start"
    outPath = getOutPath("timeLimitedStore")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("TimeLimitedStore")
    config = collections.OrderedDict()
    config["stores"] = collections.OrderedDict()
    config["refresh"] = collections.OrderedDict()
    config["slot"] = []
    config["types"] = {}
    startRowNum = 4
    h = 0
    for row in ws.rows:
        h = h + 1
        if h < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        one = collections.OrderedDict()
        if cols[0]:
            config["stores"][str(cols[0])] = one
            one["id"] = str(cols[0])
            one["name"] = str(cols[1])
            one["itemId"] = int(cols[2])
            one["count"] = int(cols[3])
            one["buyType"] = unicode(cols[4] or "")
            one["price"] = int(cols[5])
            one["levelRange"] = json.loads(cols[6])
            one["rate"] = int(cols[7])
            one["grade"] = int(cols[8])
            one["types"] = json.loads(cols[9])
            one["maxBuyCount"] = int(cols[10])
            one["des"] = unicode(cols[11] or "")

            for t in json.loads(cols[9]):
                config["types"].setdefault(str(t), [])
                config["types"][str(t)].append(one)

        if cols[13]:
            config["slot"].append({"idx": int(cols[13]), "unlockLevel": int(cols[14]), "typeList": json.loads(cols[15])})

        if cols[17]:
            config["refresh"].setdefault("auto", [])
            config["refresh"]["auto"].append(str(cols[17]))

        if cols[19]:
            config["refresh"].setdefault("price", [])
            config["refresh"]["price"].append(int(cols[19]))

    result = json.dumps(config, indent=4, ensure_ascii=False)
    result = re.sub(r"\\\\n", r"\\n", result)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "time_limited_store_config, end"


def levelRewards_config():
    """
    等级奖励配置
    """
    print "levelRewards_config, start"
    outPath = getOutPath("levelRewards")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("LevelRewards")
    config = collections.OrderedDict()
    startRowNum = 4
    h = 0
    for row in ws.rows:
        h = h + 1
        if h < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        one = collections.OrderedDict()
        if cols[0] and (cols[1] or cols[2]):
            config[str(cols[0])] = one
            one["rewards"] = json.loads(cols[1])
            one["rechargeBonus"] = int(cols[2])

    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()

    outPath = getOutPath("levelRewards", "25794.json")
    ws = wb.get_sheet_by_name("LevelRewards_25794")
    config = collections.OrderedDict()
    startRowNum = 4
    h = 0
    for row in ws.rows:
        h = h + 1
        if h < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        one = collections.OrderedDict()
        if cols[0] and (cols[1] or cols[2]):
            config[str(cols[0])] = one
            one["rewards"] = json.loads(cols[1])
            one["rechargeBonus"] = int(cols[2])

    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "levelRewards_config, end"


def level_funds_config(clientId=0):
    """
    成长基金奖励配置
    """
    sn = "LevelFunds" if clientId == 0 else "LevelFunds" + "_" + str(clientId)
    fn = "0.json" if clientId == 0 else str(clientId) + ".json"
    print "level_funds_config, start, ", sn, fn
    outPath = getOutPath("levelFunds", fn)
    wb = getWorkBook()
    ws = wb.get_sheet_by_name(sn)
    config = collections.OrderedDict()
    config["canBuyIdx"] = []
    config["funds"] = []
    config["rewards"] = collections.OrderedDict()
    startRowNum = 4
    fundIdx = 2
    rewardsIdx = 12
    h = 0
    for row in ws.rows:
        h = h + 1
        if h < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        if cols[0]:
            config["canBuyIdx"] = json.loads(cols[0])

        if cols[fundIdx]:
            one = collections.OrderedDict()
            config["funds"].append(one)
            one["productId"] = cols[fundIdx + 1]
            one["idx"] = cols[fundIdx + 0]
            one["type"] = cols[fundIdx + 2]
            one["name"] = cols[fundIdx + 3]
            one["buyType"] = cols[fundIdx + 4]
            one["price_direct"] = one["price"] = cols[fundIdx + 5]
            one["price_diamond"] = cols[fundIdx + 6]
            one["otherBuyType"] = json.loads(cols[fundIdx + 7])
            one["title"] = cols[fundIdx + 8]

        if cols[rewardsIdx]:
            config["rewards"].setdefault(str(cols[rewardsIdx + 1]), [])
            one = collections.OrderedDict()
            config["rewards"][str(cols[rewardsIdx + 1])].append(one)
            one["level"] = int(cols[rewardsIdx + 0])
            one["free_rewards"] = json.loads(cols[rewardsIdx + 3])
            one["rechargeBonus"] = int(cols[rewardsIdx + 4])
            one["funds_rewards"] = json.loads(cols[rewardsIdx + 6])

    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "level_funds_config, end"


def super_egg_config(clientId=0):
    """
    超级扭蛋活动配置
    """
    print "super_egg_config, start, "
    outPath = getOutPath("superEggs")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("SuperEggs")
    config = collections.OrderedDict()
    startRowNum = 4
    eggIdx = 2
    h = 0
    for row in ws.rows:
        h = h + 1
        if h < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        if cols[0]:
            config["maxBuyCount"] = int(cols[0])

        if cols[eggIdx]:
            one = collections.OrderedDict()
            config.setdefault("eggs", [])
            config["eggs"].append(one)
            one["productId"] = cols[eggIdx]
            one["name"] = cols[eggIdx + 1]
            one["des"] = cols[eggIdx + 2]
            one["buyType"] = cols[eggIdx + 3]
            one["price_direct"] = cols[eggIdx + 4]
            one["price_diamond"] = cols[eggIdx + 5]
            one["dropConf"] = []
            for i in range(5):
                one["dropConf"].append({"rewards": json.loads(cols[eggIdx + 7 + 3 * i]), "color": cols[eggIdx + 7 + 3 * i + 1], "rate": cols[eggIdx + 7 + 3 * i + 2]})
    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "super_egg_config, end"


def updateVerRewards_config():
    """
    更服奖励配置
    """
    print "updateVerRewards_config, start"
    outPath = getOutPath("updateVerRewards")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("UpdateVerRewards")
    config = collections.OrderedDict()
    startRowNum = 4
    h = 0
    for row in ws.rows:
        h = h + 1
        if h < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        one = collections.OrderedDict()
        config.setdefault("rewards", {})
        if cols[0] is not None:
            config["rewards"][str(cols[0])] = one
            one["vip"] = int(cols[0])
            one["1"] = json.loads(cols[1])
            one["2"] = json.loads(cols[2])

        if cols[4]:
            config["version"] = str(cols[4])
        if cols[5]:
            config["type"] = int(cols[5])
        if cols[6]:
            config["msg"] = unicode(cols[6] or "")


    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "updateVerRewards_config, end"


def checkin_config():
    """
    更服奖励配置
    """
    print "checkin_config, start"
    outPath = getOutPath("checkin")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("Checkin")
    config = collections.OrderedDict()
    startRowNum = 4
    h = 0
    for row in ws.rows:
        h = h + 1
        if h < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        if cols[0] is not None:
            config.setdefault("oldcheckinData", collections.OrderedDict())
            config["oldcheckinData"][str(cols[0])] = {"normalReward": json.loads(cols[1]), "shareReward": json.loads(cols[2])}

        if cols[4] is not None:
            config.setdefault("resetInfo", collections.OrderedDict())
            config["resetInfo"]["vip"] = int(cols[4])
            config["resetInfo"]["resetWeekDay"] = int(cols[5])

        if cols[8] is not None:
            config.setdefault("checkinData", collections.OrderedDict())
            one = collections.OrderedDict()
            config["checkinData"][str(cols[8])] = one
            one["type"] = str(cols[8])
            one["unlockdays"] = int(cols[9])
            one["datas"] = []
            for x in range(10, len(cols), 2):
                if cols[x]:
                    if cols[8] != "multiple":
                        one["datas"].append({"rewards": json.loads(cols[x]), "rate": int(cols[x + 1])})
                    else:
                        one["datas"].append({"rewards": int(cols[x]), "rate": int(cols[x + 1])})

    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "checkin_config, end"


def prizewheel_config():
    """
    渔场轮盘配置
    """
    print "prizewheel_config, start"
    outPath = getOutPath("prizeWheel")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("PrizeWheel")
    config = collections.OrderedDict()
    config["maxSpinTimes"] = 1
    config["energy"] = []
    config["prize"] = collections.OrderedDict()
    config["bet"] = collections.OrderedDict()
    startRowNum = 4
    h = 0

    wheelIdx = 16
    betIdx = 51
    for row in ws.rows:
        h = h + 1
        if h < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        if cols[0]:
            config["maxSpinTimes"] = int(cols[0])

        if cols[2]:
            one = collections.OrderedDict()
            config["energy"].append(one)
            for i in range(3, 14, 3):
                one[str(cols[i])] = collections.OrderedDict()
                one[str(cols[i])]["fishPool"] = int(cols[i])
                one[str(cols[i])]["cost"] = int(cols[i + 1])
                one[str(cols[i])]["add"] = json.loads(cols[i + 2])

        if cols[wheelIdx]:
            config["prize"].setdefault(str(cols[wheelIdx]), collections.OrderedDict())
            config["prize"][str(cols[wheelIdx])]["nextRoomId"] = int(cols[wheelIdx + 1])
            config["prize"][str(cols[wheelIdx])]["nextRoomMultiple"] = int(cols[wheelIdx + 2])
            config["prize"][str(cols[wheelIdx])]["betList"] = json.loads(cols[wheelIdx + 3])
            wheel = []
            for i in range(wheelIdx + 4, wheelIdx + 34, 3):
                one = collections.OrderedDict()
                one["rewards"] = json.loads(cols[i])
                one["rate"] = int(cols[i + 1])
                one["reset"] = int(cols[i + 2])
                wheel.append(one)
            config["prize"][str(cols[wheelIdx])]["wheel"] = wheel

        if cols[betIdx]:
            one = collections.OrderedDict()
            config["bet"][str(cols[betIdx])] = one
            for i in range(betIdx + 1, betIdx + 2, 2):
                one[str(cols[i])] = json.loads(cols[i + 1])

    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "prizewheel_config, end"


def grandPrixPrizewheel_config():
    """
    渔场轮盘配置
    """
    print "grandPrixPrizewheel_config, start"
    outPath = getOutPath("grandPrixPrizeWheel")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("GrandPrixPrizeWheel")
    config = collections.OrderedDict()
    config["maxSpinTimes"] = 1
    config["ratio"] = {}
    config["energy"] = []
    config["prize"] = collections.OrderedDict()
    config["bet"] = collections.OrderedDict()
    startRowNum = 4
    h = 0

    wheelIdx = 18
    betIdx = 51
    for row in ws.rows:
        h = h + 1
        if h < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        if cols[0]:
            config["maxSpinTimes"] = int(cols[0])

        if cols[2]:
            config["ratio"] = {"fire": float(cols[2]), "loss": float(cols[3])}      # 开火充能系数 亏损充能系数

        if cols[5]:
            one = collections.OrderedDict()
            config["energy"].append(one)
            for i in range(5, 16, 3):
                one[str(cols[i])] = collections.OrderedDict()
                one[str(cols[i])]["fpMultiple"] = int(cols[i])                      # 倍率
                one[str(cols[i])]["addUnit"] = int(cols[i + 1])                     # 单次充能
                one[str(cols[i])]["cost"] = int(cols[i + 2])                        # 单次耗能
                one[str(cols[i])]["add"] = {}                                       #

        if cols[wheelIdx]:
            config["prize"].setdefault(str(cols[wheelIdx]), collections.OrderedDict())
            config["prize"][str(cols[wheelIdx])]["nextRoomId"] = 0
            config["prize"][str(cols[wheelIdx])]["nextRoomMultiple"] = 0
            config["prize"][str(cols[wheelIdx])]["betList"] = json.loads(cols[wheelIdx + 1])
            wheel = []
            for i in range(wheelIdx + 2, wheelIdx + 32, 3):
                one = collections.OrderedDict()
                one["rewards"] = json.loads(cols[i])
                one["rate"] = int(cols[i + 1])
                one["reset"] = int(cols[i + 2])
                wheel.append(one)
            config["prize"][str(cols[wheelIdx])]["wheel"] = wheel

        if cols[betIdx]:
            one = collections.OrderedDict()
            config["bet"][str(cols[betIdx])] = one
            for i in range(betIdx + 1, betIdx + 2, 2):
                one[str(cols[i])] = json.loads(cols[i + 1])

    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "grandPrixPrizewheel_config, end"


def statistic_prop_config():
    """
    BI需要统计的道具配置
    """
    print "statistic_prop_config, start"
    reload(sys)
    sys.setdefaultencoding("utf-8")
    outPath = getOutPath("statisprop")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("StatisProp")
    config = collections.OrderedDict()
    config["statisprop"] = []
    config["statisPropId"] = []
    startRowNum = 4
    i = 0
    for row in ws.rows:
        i = i+1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if cols[0]:
            one = collections.OrderedDict()
            config["statisprop"].append(one)
            one["propid"] = int(cols[0])
            one["propname"] = str(cols[1])
        if cols[0]:
            config["statisPropId"].append(int(cols[0]))
    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "statistic_prop_config, end"


def slot_machine_config():
    """
    老虎机活动配置
    """
    reload(sys)
    sys.setdefaultencoding("utf-8")
    print "slot_machine_config, start"
    outPath = getOutPath("slotMachine")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("SlotMachine")
    config = collections.OrderedDict()
    config["slotmachine"] = []
    config["integralrewards"] = []
    config["freetime"] = []
    config["maxPlayTimes"] = -1
    config["outProductId"] = []

    startRowNum = 4
    h = 0
    for row in ws.rows:
        h = h + 1
        if h < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        if cols[0]:
            one = collections.OrderedDict()
            config["slotmachine"].append(one)
            one["weight"] = int(cols[1])
            one["rewards"] = []

            item = {}
            item["rewardIdx"] = int(cols[0])
            item["name"] = int(cols[2])
            item["count"] = int(cols[3])
            item["rare"] = int(cols[4])
            item["des"] = str(cols[5])
            item["integral"] = int(cols[6])
            one["rewards"].append(item)

        if cols[8]:
            one = collections.OrderedDict()
            config["integralrewards"].append(one)
            one["rank"] = int(cols[8])
            one["rewards"] = []
            item = {}
            item["name"] = int(cols[9])
            item["count"] = int(cols[10])
            item["des"] = str(cols[11])
            one["rewards"].append(item)

        if cols[13] is not None:
            config["freetime"].append(int(cols[13]))

        if cols[15]:
            config["maxPlayTimes"] = int(cols[15])

        if cols[17]:
            config["outProductId"].append(str(cols[17]))

    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "slot_machine_config, end"


def money_tree_config():
    """
    摇钱树活动数值配置
    """
    reload(sys)
    sys.setdefaultencoding("utf-8")
    print "money_tree_config, start"
    outPath = getOutPath("moneyTree")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("MoneyTree")
    config = collections.OrderedDict()
    config["consumeCount"] = collections.OrderedDict()
    config["consumeItem"] = []
    config["moneyTree"] = []
    config["freeTime"] = []
    config["outProductId"] = []
    config["rule"] = ""

    startRowNum = 4
    h = 0
    for row in ws.rows:
        h = h + 1
        if h < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        if cols[0] is not None:
            one = collections.OrderedDict()
            config["consumeCount"][str(cols[0])] = one
            one["vipLevel"] = int(cols[0])
            one["count"] = int(cols[1])
        if cols[3] is not None:
            one = collections.OrderedDict()
            config["moneyTree"].append(one)
            one["weight"] = int(cols[4])
            item = collections.OrderedDict()
            item["rewardIdx"] = int(cols[3])
            item["name"] = int(cols[5])
            item["count"] = int(cols[6])
            item["rare"] = int(cols[7])
            item["display"] = int(cols[8])
            item["desc"] = str(cols[9])
            one["reward"] = item
            one["certainCount"] = int(cols[10])
            one["drawCount"] = int(cols[11])
            one["vipLevel"] = int(cols[12])
            one["level"] = int(cols[13])
            one["obtainCount"] = int(cols[14])
            one["serverCount"] = int(cols[15])
        if cols[17] is not None:
            config["consumeItem"] = json.loads(cols[17])
        if cols[19] is not None:
            config["freeTime"].append(int(cols[19]))
        if cols[21] is not None:
            config["outProductId"].append(str(cols[21]))
        if cols[23] is not None:
            config["rule"] = str(cols[23])

    result = json.dumps(config, indent=4, ensure_ascii=False)
    result = re.sub(r"\\\\n", r"\\n", result)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "money_tree_config, end"


def canned_fish_config():
    """
    鱼罐厂配置
    """
    print "canned_fish_config, start"
    reload(sys)
    sys.setdefaultencoding("utf-8")
    outPath = getOutPath("cannedFish")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("CannedFish")
    config = collections.OrderedDict()
    config["welfareCan"] = collections.OrderedDict()
    config["rule"] = ""
    config["makeMail"] = ""
    config["receiveMail"] = ""
    config["bagMail"] = ""
    config["canProp"] = []
    config["wordFactoryChip"] = None
    config["userChip"] = None
    config["expireTime"] = None
    config["userMakeCan"] = collections.OrderedDict()
    config["welfareCanGenerateTs"] = None
    config["welfareCanExpireTs"] = None
    config["maxCount"] = None


    startRowNum = 4
    h = 0
    for row in ws.rows:
        h = h + 1
        if h < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        if cols[0]:
            one = collections.OrderedDict()
            config["welfareCan"][str(cols[0])] = one
            one["name"] = int(cols[1])
            one["count"] = int(cols[2])
            one["des"] = str(cols[3])

        if cols[5] is not None:
            config["vipLimit"] = int(cols[5])

        if cols[7] is not None:
            config["rule"] = str(cols[7])

        if cols[9] is not None:
            config["makeMail"] = str(cols[9])

        if cols[11] is not None:
            config["receiveMail"] = str(cols[11])

        if cols[13] is not None:
            config["bagMail"] = str(cols[13])

        if cols[15] is not None:
            config["canProp"].append(int(cols[15]))

        if cols[18]:
            config["wordFactoryChip"] = int(cols[18])

        if cols[19]:
            config["userChip"] = int(cols[19])

        if cols[21]:
            config["expireTime"] = float(cols[21])

        if cols[23]:
            one = collections.OrderedDict()
            config["userMakeCan"][str(cols[23])] = one
            one["name"] = int(cols[23])
            one["des"] = str(cols[24])

        if cols[26]:
            config["welfareCanGenerateTs"] = int(cols[26])

        if cols[27]:
            config["welfareCanExpireTs"] = int(cols[27])

        if cols[29]:
            config["maxCount"] = int(cols[29])


    result = json.dumps(config, indent=4, ensure_ascii=False)
    result = re.sub(r"\\\\n", r"\\n", result)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "canned_fish_config, end"


def credit_store_config():
    """
    微信公众号积分商城配置
    """
    reload(sys)
    sys.setdefaultencoding("utf-8")
    print "credit_store_config, start"
    outPath = getOutPath("creditStore")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("CreditStore")
    config = collections.OrderedDict()
    config["creditStore"] = []
    config["creditTask"] = []

    startRowNum = 4
    h = 0
    for row in ws.rows:
        h = h + 1
        if h < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        if cols[0] is not None:
            one = collections.OrderedDict()
            one["productId"] = str(cols[0])
            one["name"] = str(cols[1])
            one["kindId"] = int(cols[2])
            one["count"] = int(cols[3])
            one["credit"] = int(cols[4])
            one["limitType"] = int(cols[5])
            one["limitCount"] = int(cols[6])
            one["limitVipLevel"] = int(cols[7])
            if cols[8]:
                one["popInfo"] = str(cols[8])
            one["pic"] = str(cols[9])
            config["creditStore"].append(one)
        else:
            break

        if cols[11] is not None:
            one = collections.OrderedDict()
            one["name"] = str(cols[11])
            one["desc"] = str(cols[12])
            one["pic"] = str(cols[13])
            config["creditTask"].append(one)

    result = json.dumps(config, indent=4, ensure_ascii=False)
    result = re.sub(r"\\\\n", r"\\n", result)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "credit_store_config, end"


def supply_box_config():
    """
    补给箱活动配置
    """
    reload(sys)
    sys.setdefaultencoding("utf-8")
    print "supply_box_config, start"
    outPath = getOutPath("supplyBox")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("SupplyBox")
    config = collections.OrderedDict()
    config["supplyBox"] = []
    config["boxes"] = collections.OrderedDict()
    config["timeDesc"] = ""
    config["rule"] = ""
    config["actType"] = 0

    startRowNum = 4
    h = 0
    for row in ws.rows:
        h = h + 1
        if h < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        if cols[0] is not None:
            one = collections.OrderedDict()
            one["boxId"] = int(cols[0])
            one["boxName"] = str(cols[1])
            one["star"] = int(cols[2])
            one["taskIds"] = json.loads(cols[3])
            config["supplyBox"].append(one)

        if cols[5] is not None:
            config["boxes"].setdefault(int(cols[5]), [])
            one = collections.OrderedDict()
            reward = collections.OrderedDict()
            one["weight"] = int(cols[6])
            one["reward"] = reward
            reward["name"] = int(cols[7])
            reward["count"] = int(cols[8])
            reward["rare"] = int(cols[9])
            reward["desc"] = str(cols[10]) if cols[10] else ""
            one["vipLevel"] = int(cols[11])
            config["boxes"][int(cols[5])].append(one)

        if cols[13] is not None:
            config["timeDesc"] = str(cols[13])
        if cols[15] is not None:
            config["rule"] = str(cols[15])
        if cols[17] is not None:
            config["actType"] = int(cols[17])

    result = json.dumps(config, indent=4, ensure_ascii=False)
    result = re.sub(r"\\\\n", r"\\n", result)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "supply_box_config, end"


def bigPrize_config():
    """
    幸运降临配置
    """
    print "bigprize_config, start"
    outPath = getOutPath("bigPrize")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("BigPrize")
    config = collections.OrderedDict()

    startRowNum = 4
    h = 0
    for row in ws.rows:
        h = h + 1
        if h < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        if cols[0]:
            config[str(cols[0])] = cols[1]
    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "bigprize_config, end"


def festival_turntable_config():
    """
    节日转盘抽大奖活动配置(万圣节,感恩节)
    """
    reload(sys)
    sys.setdefaultencoding("utf-8")
    print "festival_turntable_config, start"
    outPath = getOutPath("festivalTurntable")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("FestivalTurntable")
    config = collections.OrderedDict()
    config["festivalTurntable"] = []
    config["integralrewards"] = []
    config["consumeItmeId"] = None
    config["rankName"] = None
    config["itemTxt"] = None
    config["jumpToactivity"] = None

    startRowNum = 4
    h = 0
    for row in ws.rows:
        h = h + 1
        if h < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        if cols[0]:
            one = collections.OrderedDict()
            config["festivalTurntable"].append(one)
            one["weight"] = int(cols[1])
            one["rewards"] = []

            item = {}
            item["rewardIdx"] = int(cols[0])
            item["name"] = int(cols[2])
            item["count"] = int(cols[3])
            item["rare"] = int(cols[4])
            item["des"] = str(cols[5])
            item["integral"] = int(cols[6])
            one["rewards"].append(item)

        if cols[8]:
            one = collections.OrderedDict()
            config["integralrewards"].append(one)
            one["rank"] = int(cols[8])
            one["rewards"] = []
            item = {}
            item["name"] = int(cols[9])
            item["count"] = int(cols[10])
            item["des"] = str(cols[11])
            one["rewards"].append(item)

        if cols[13]:
            config["consumeItmeId"] = int(cols[13])
        if cols[15]:
            config["rankName"] = str(cols[15])
        if cols[17]:
            config["itemTxt"] = str(cols[17])
        if cols[19] is not None:
            config["jumpToactivity"] = str(cols[19])
            if config["jumpToactivity"] == "0":
                config["jumpToactivity"] = int(cols[19])

    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "festival_turntable_config, end"


def store_config(clientId=0):
    reload(sys)
    sys.setdefaultencoding("utf-8")

    sn = "weixin" if clientId == 0 else str(clientId)
    fn = "0.json" if clientId == 0 else str(clientId) + ".json"
    outPath = getOutPath("gift", fn)
    outPath = getOutPath("store", fn)
    ws = getWorkBook("store.xlsx").get_sheet_by_name(sn)
    config = collections.OrderedDict()
    config["coinStore"] = collections.OrderedDict()
    config["diamondStore"] = collections.OrderedDict()
    config["pearlStore"] = collections.OrderedDict()
    config["chestStore"] = collections.OrderedDict()
    config["chestStore"]["items"] = collections.OrderedDict()
    config["chestStore"]["ads"] = collections.OrderedDict()
    config["couponStore"] = collections.OrderedDict()
    config["gunSkinStore"] = collections.OrderedDict()
    config["bulletStore"] = collections.OrderedDict()
    config["bulletStore"]["hall37"] = collections.OrderedDict()

    i = 0 
    results = []
    startRowNum = 4
    diamond_store_num = 15
    pearl_store_num = 35
    chest_store_num = 50
    ads_num = 71
    coupon_store_num = 73
    gun_skin_num = 96
    robbery_store_num = 110

    for row in ws.rows:
        i = i+1
        if i < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        if cols[0] is not None:
            one = collections.OrderedDict()
            config["coinStore"][str(cols[0])] = one
            one["name"] = json.loads(cols[1])
            one["itemId"] = int(cols[2])
            one["count"] = json.loads(cols[3])
            one["label"] = json.loads(cols[4])
            one["order"] = int(cols[5])
            one["price"] = int(cols[6])
            one["price_direct"] = int(cols[7])
            one["price_diamond"] = int(cols[8])
            one["additionVip"] = int(cols[9])
            one["tag"] = int(cols[10])
            one["addition"] = json.loads(cols[11])
            if cols[12] == "\"\"":
                one["pic"] = ""
            else:
                one["pic"] = str(cols[12])
            one["buyType"] = str(cols[13])

        if cols[diamond_store_num]:
            one = collections.OrderedDict()
            config["diamondStore"][str(cols[diamond_store_num])] = one
            one["type"] = str(cols[diamond_store_num+1])
            one["name"] = json.loads(cols[diamond_store_num+2])
            one["itemId"] = int(cols[diamond_store_num+3])
            one["count"] = json.loads(cols[diamond_store_num+4])
            one["label"] = json.loads(cols[diamond_store_num+5])
            one["order"] = int(cols[diamond_store_num+6])
            one["price"] = int(cols[diamond_store_num+7])
            one["price_direct"] = int(cols[diamond_store_num+8])
            one["price_diamond"] = int(cols[diamond_store_num+9])
            one["additionVip"] = int(cols[diamond_store_num+10])
            one["tag"] = int(cols[diamond_store_num+11])
            one["addition"] = json.loads(cols[diamond_store_num+12])
            if cols[diamond_store_num+13] == "\"\"":
                one["pic"] = ""
            else:
                one["pic"] = str(cols[diamond_store_num+13])
            one["buyType"] = str(cols[diamond_store_num+14])
            if cols[diamond_store_num+15]:
                one["otherBuyType"] = json.loads(cols[diamond_store_num+15])
            if cols[diamond_store_num+16] is not None:
                one["visible"] = int(cols[diamond_store_num+16])
            if cols[diamond_store_num+17] is not None:
                one["appVisible"] = int(cols[diamond_store_num+17])
            #one["extend"] = json.loads(cols[30])

        if cols[pearl_store_num]:
            one = collections.OrderedDict()
            config["pearlStore"][str(cols[pearl_store_num])] = one
            one["name"] = json.loads(cols[pearl_store_num+1])
            one["itemId"] = int(cols[pearl_store_num+2])
            one["count"] = json.loads(cols[pearl_store_num+3])
            one["order"] = int(cols[pearl_store_num+4])
            one["price"] = int(cols[pearl_store_num+5])
            one["price_direct"] = int(cols[pearl_store_num+6])
            one["price_diamond"] = int(cols[pearl_store_num+7])
            one["additionVip"] = int(cols[pearl_store_num+8])
            one["tag"] = int(cols[pearl_store_num+9])
            one["addition"] = json.loads(cols[pearl_store_num+10])
            if cols[pearl_store_num+11] == "\"\"":
                one["pic"] = ""
            else:
                one["pic"] = str(cols[pearl_store_num+11])
            one["buyType"] = str(cols[pearl_store_num+12])
            one["otherBuyType"] = json.loads(cols[pearl_store_num+13])

        if cols[chest_store_num]:
            one = collections.OrderedDict()
            config["chestStore"]["items"][str(cols[chest_store_num])] = one
            one["name"] = str(cols[chest_store_num+1])
            one["itemId"] = int(cols[chest_store_num+2])
            if cols[chest_store_num+3]:
                one["type"] = int(cols[chest_store_num+3])
            if str(cols[chest_store_num+4]).find(",") < 0:
                one["count"] = int(cols[chest_store_num+4])
            else:
                one["count"] = json.loads(cols[chest_store_num+4])
            one["order"] = int(cols[chest_store_num+5])
            one["price"] = int(cols[chest_store_num+6])
            one["discountPrice"] = json.loads(cols[chest_store_num+7])
            if cols[chest_store_num+8]:
                one["price_direct"] = int(cols[chest_store_num+8])
            if cols[chest_store_num+9]:
                one["price_diamond"] = int(cols[chest_store_num+9])
            one["tag"] = json.loads(cols[chest_store_num+10])
            if cols[chest_store_num+11] == "\"\"":
                one["desc"] = ""
            else:
                one["desc"] = str(cols[chest_store_num+11])
            one["addition"] = json.loads(cols[chest_store_num+12])
            one["buyType"] = str(cols[chest_store_num+13])
            if cols[chest_store_num+14]:
                one["otherBuyType"] = json.loads(cols[chest_store_num+14])
            if cols[chest_store_num+15]:
                one["minLevel"] = int(cols[chest_store_num+15])
            if cols[chest_store_num+16]:
                one["lowVersion"] = str(cols[chest_store_num+16])
            if cols[chest_store_num+17]:
                one["convenientBuy"] = int(cols[chest_store_num+17])
            if cols[chest_store_num+18]:
                one["additionVip"] = int(cols[chest_store_num+18])
            if cols[chest_store_num+19]:
                one["extend"] = json.loads(cols[chest_store_num+19])
            if cols[chest_store_num+20]:
                one["activityPrice"] = json.loads(cols[chest_store_num+20])
            if cols[ads_num]:
                config["chestStore"]["ads"] = json.loads(cols[ads_num])
                #one["chestStore"]["ads"] = json.loads(cols[ads_num])

        if cols[coupon_store_num]:
            one = collections.OrderedDict()
            config["couponStore"][str(cols[coupon_store_num])] = one
            one["name"] = str(cols[coupon_store_num + 1])
            one["itemId"] = int(cols[coupon_store_num + 2])
            one["count"] = int(cols[coupon_store_num + 3])
            one["order"] = int(cols[coupon_store_num + 4])
            if cols[coupon_store_num + 5] != "\"\"":
                one["desc"] = str(cols[coupon_store_num + 5])
            else:
                one["desc"] = ""
            one["price"] = int(cols[coupon_store_num + 6])
            one["price_coupon"] = int(cols[coupon_store_num + 7])
            if cols[coupon_store_num+8]:
                one["activityPrice"] = json.loads(cols[coupon_store_num+8])
            one["tag"] = int(cols[coupon_store_num + 9])
            one["addition"] = json.loads(cols[coupon_store_num + 10])
            if cols[coupon_store_num + 11] != "\"\"":
                one["pic"] = str(cols[coupon_store_num + 11])
            else:
                one["pic"] = ""
            one["guideLimit"] = int(cols[coupon_store_num + 12])
            if cols[coupon_store_num+13]:
                one["dayBuyNum"] = int(cols[coupon_store_num+13])
            one["serverBuyNum"] = int(cols[coupon_store_num + 14]) if cols[coupon_store_num + 14] else 0
            one["vip"] = int(cols[coupon_store_num + 15]) if cols[coupon_store_num + 15] else 0
            if cols[coupon_store_num+16] is not None:
               one["visible"] = cols[coupon_store_num+16]
            one["buyType"] = str(cols[coupon_store_num + 17])
            if cols[coupon_store_num+18] is not None:
                one["reviewVerLimit"] = int(cols[coupon_store_num+18])
            if cols[coupon_store_num+19] is not None:
                one["entityItem"] = int(cols[coupon_store_num+19])
            if cols[coupon_store_num+20]:
                one["displayLimitNum"] = int(cols[coupon_store_num+20])
            if cols[coupon_store_num + 21]:
                one["extend"] = json.loads(cols[coupon_store_num + 21])

        if cols[gun_skin_num]:
            one = collections.OrderedDict()
            config["gunSkinStore"][str(cols[gun_skin_num])] = one
            one["name"] = str(cols[gun_skin_num+1])
            one["itemId"] = int(cols[gun_skin_num+2])
            one["count"] = int(cols[gun_skin_num+3])
            one["order"] = int(cols[gun_skin_num+4])
            if cols[gun_skin_num+5] is not None:
                one["vip"] = int(cols[gun_skin_num+5])
            if cols[gun_skin_num+6] != "\"\"":
                one["desc"] = str(cols[gun_skin_num+6])
            else:
                one["desc"] = ""
            one["price"] = int(cols[gun_skin_num+7])
            one["discountPrice"] = json.loads(cols[gun_skin_num+8])
            one["tag"] = int(cols[gun_skin_num+9])
            one["addition"] = json.loads(cols[gun_skin_num+10])
            if cols[gun_skin_num+11] != "\"\"":
                one["pic"] = str(cols[gun_skin_num+11])
            else:
                one["pic"] = ""
            one["buyType"] = str(cols[gun_skin_num+12])

        if cols[robbery_store_num]:
            one = collections.OrderedDict()
            config["bulletStore"]["hall37"][str(cols[robbery_store_num])] = one
            one["name"] = str(cols[robbery_store_num+1])
            one["itemId"] = int(cols[robbery_store_num+2])
            one["count"] = int(cols[robbery_store_num+3])
            one["order"] = int(cols[robbery_store_num+4])
            one["price"] = int(cols[robbery_store_num+5])
            one["price_direct"] = int(cols[robbery_store_num+6])
            one["price_diamond"] = int(cols[robbery_store_num+7])
            one["tag"] = int(cols[robbery_store_num+8])
            one["addition"] = json.loads(cols[robbery_store_num+9])
            if cols[robbery_store_num+10] != "\"\"":
                one["pic"] = str(cols[robbery_store_num+10])
            else:
                one["pic"] = ""
            one["buyType"] = str(cols[robbery_store_num+11])
            one["robberyBonus"] = int(cols[robbery_store_num+12])
            one["vipAddition"] = json.loads(cols[robbery_store_num+13])

    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "store_config, end"


def collect_item_config():
    """
    收集道具赢奖活动配置(赢永久魅影皮肤活动)
    """
    reload(sys)
    sys.setdefaultencoding("utf-8")
    print "collect_item_config, start"
    outPath = getOutPath("collectItem")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("CollectItem")
    config = collections.OrderedDict()
    config["consumeItemId"] = None
    config["rankIdx"] = None
    config["ts"] = None
    config["activityTxt"] = None
    config["skinId"] = None

    startRowNum = 4
    h = 0
    for row in ws.rows:
        h = h + 1
        if h < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        if cols[0]:
            config["consumeItemId"] = int(cols[0])

        if cols[2]:
            config["rankIdx"] = int(cols[2])
        if cols[4]:
            config["ts"] = float(cols[4])
        if cols[6]:
            config["activityTxt"] = str(cols[6])
        if cols[8]:
            config["skinId"] = int(cols[8])
            
    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "collect_item_config, end"


def compAct_config():
    """
    竞赛活动配置
    """
    print "compAct_config, start"
    outPath = getOutPath("competition")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("Competition")
    config = collections.OrderedDict()
    config["items"] = []
    config["points"] = collections.OrderedDict()
    config["timeRange"] = None
    config["bonus"] = None
    config["winnerRankRewards"] = []
    config["rankRewards"] = []
    config["ratio"] = None
    config["balance"] = None

    startRowNum = 4
    timeRangeCols = 3
    bonusCols = 5
    itemsCols = 9
    balanceCols = 29
    h = 0
    for row in ws.rows:
        h = h + 1
        if h < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        if cols[0]:
            config["points"].setdefault(str(cols[0]), int(cols[1]))
        if cols[timeRangeCols]:
            config["timeRange"] = json.loads(cols[timeRangeCols])
        if cols[bonusCols]:
            config["bonus"] = {"maxBonus": int(cols[bonusCols]),
                               "minBonus": int(cols[bonusCols + 1]),
                               "minCount": int(cols[bonusCols + 2])}
        if cols[itemsCols]:
            one = collections.OrderedDict()
            config["items"].append(one)
            one["productId"] = str(cols[itemsCols])
            one["name"] = str(cols[itemsCols + 1])
            one["buyType"] = str(cols[itemsCols + 2])
            one["price_direct"] = int(cols[itemsCols + 3])
            one["price_diamond"] = int(cols[itemsCols + 4])
            one["itemId"] = int(cols[itemsCols + 5])
            one["count"] = int(cols[itemsCols + 6])
            one["extraCount"] = int(cols[itemsCols + 7])
            one["inspireTime"] = int(cols[itemsCols + 8])
            one["extraGainTime"] = int(cols[itemsCols + 9])

        if cols[itemsCols + 11]:
            config["winnerRankRewards"].append(json.loads(cols[itemsCols + 11]))
        if cols[itemsCols + 12]:
            config["rankRewards"].append(json.loads(cols[itemsCols + 12]))
        if cols[itemsCols + 13]:
            config["rankRewardsTeamCnt"] = int(cols[itemsCols + 13])

        if cols[itemsCols + 15]:
            config["ratio"] = json.loads(cols[itemsCols + 15])

        if cols[balanceCols]:
            one = collections.OrderedDict()
            config["balance"] = one
            one["enable"] = int(cols[balanceCols])
            one["time"] = json.loads(cols[balanceCols + 1])
            one["point"] = json.loads(cols[balanceCols + 2])
            one["diff"] = int(cols[balanceCols + 3])
            one["interval"] = int(cols[balanceCols + 4])

    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "compAct_config, end"


def _getSheetName(confName, clientId=0):
    return confName if clientId == 0 else confName + "_" + str(clientId)


def _getFileName(clientId=0):
    return "0.json" if clientId == 0 else str(clientId) + ".json"


def newbie7DaysGift_config(clientId=0):
    """
    新手7日礼包配置
    """
    sn = _getSheetName("Newbie7DaysGift", clientId)
    fn = _getFileName(clientId)
    print "newbie7DaysGift_config, start, ", sn, fn
    outPath = getOutPath("newbie7DaysGift", fn)
    wb = getWorkBook()
    ws = wb.get_sheet_by_name(sn)
    config = []
    startRowNum = 4
    h = 0
    for row in ws.rows:
        h = h + 1
        if h < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        if cols[1]:
            one = collections.OrderedDict()
            config.append(one)
            one["idx"] = int(cols[0])
            one["rewards"] = json.loads(cols[1])
            one["cond"] = json.loads(cols[2])
            one["des"] = str(cols[3])
            one["rechargeBonus"] = int(cols[4])

    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "newbie7DaysGift_config, end"


def lottery_ticket_config(clientId=0):
    """
    红包券抽奖配置
    """
    sn = "LotteryTicket" if clientId == 0 else "LotteryTicket" + "_" + str(clientId)
    fn = "0.json" if clientId == 0 else str(clientId) + ".json"
    print "lottery_ticket_config, start, ", sn, fn
    outPath = getOutPath("lotteryTicket", fn)
    wb = getWorkBook()
    ws = wb.get_sheet_by_name(sn)
    config = collections.OrderedDict()
    config["showPrize"] = collections.OrderedDict()
    config["rewards"] = collections.OrderedDict()
    config["startTime"] = ""
    config["expire"] = None
    config["exchangeReward"] = []
    config["exchangeRewardIgnore"] = []
    config["rule"] = ""
    config["ruleIgnore"] = ""
    config["lowClientVersion"] = None


    startRowNum = 4
    effectiveTimeCol = 13

    h = 0
    for row in ws.rows:
        h = h + 1
        if h < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        if cols[0]:
            one = collections.OrderedDict()
            config["showPrize"][str(cols[0])] = one
            one["prize"] = json.loads(cols[2])
            one["roomMultiple"] = int(cols[1])
            one["maxPrize"] = int(cols[3])
            one["nextRoomId"] = int(cols[4])

        if cols[6]:
            one = collections.OrderedDict()
            config["rewards"][str(cols[6])] = one
            one["roomId"] = int(cols[7])
            one["costChipList"] = json.loads(cols[8])
            one["rewardList"] = json.loads(cols[9])
            one["totalDrawCount"] = int(cols[10])
            one["orderIdx"] = int(cols[11])

        if cols[effectiveTimeCol]:
            config["startTime"] = str(cols[effectiveTimeCol])
        if cols[14]:
            config["expire"] = int(cols[14])
        if cols[16]:
            config["exchangeReward"] = json.loads(cols[16])
        if cols[17]:
            config["exchangeRewardIgnore"] = json.loads(cols[17])
        if cols[19]:
            config["rule"] = str(cols[19])
        if cols[20]:
            config["ruleIgnore"] = str(cols[20])
        if cols[22]:
            config["lowClientVersion"] = (cols[22])


    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "lottery_ticket_config, end"


def pass_card_config(clientId=0):
    """
    通行证活动配置
    """
    print "pass_card_config, start, "
    outPath = getOutPath("passCard")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("PassCard")
    config = collections.OrderedDict()
    config["productInfo"] = collections.OrderedDict()
    config["productInfoAcId"] = collections.OrderedDict()
    startRowNum = 4
    h = 0
    for row in ws.rows:
        h = h + 1
        if h < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)

        if cols[0]:
            one = collections.OrderedDict()
            config["productInfo"][str(cols[0])] = one
            one["productId"] = str(cols[1])
            one["productName"] = str(cols[2])
            one["buyType"] = str(cols[3])
            one["price_direct"] = int(cols[4])
            one["price_diamond"] = int(cols[5])
            one["productDesc"] = str(cols[6])

        if cols[0]:
            one = collections.OrderedDict()
            config["productInfoAcId"][str(cols[1])] = one
            one["acId"] = str(cols[0])
            one["buyType"] = str(cols[3])
            one["price_direct"] = int(cols[4])
            one["price_diamond"] = int(cols[5])

    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "pass_card_config, end"


def skill_compen_config():
    """
    技能补偿配置
    """
    print "skill_compen_config, start, "
    outPath = getOutPath("skillCompensate")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("SkillCompensate")
    config = collections.OrderedDict()
    config["skillList"] = []
    config["skillLevComp"] = collections.OrderedDict()
    config["starLevComp"] = collections.OrderedDict()
    config["gradeItemId"] = None
    config["starItemId"] = None
    startRowNum = 4
    h = 0
    for row in ws.rows:
        h = h + 1
        if h < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        if cols[0]:
            config["skillList"].append(int(cols[0]))
            config["gradeItemId"] = int(cols[1])
            config["starItemId"] = int(cols[2])

        if cols[4]:
            one = collections.OrderedDict()
            config["skillLevComp"][str(cols[4])] = int(cols[5])
        if cols[7]:
            one = collections.OrderedDict()
            config["starLevComp"][str(cols[7])] = int(cols[8])
        if cols[10]:
            config["compenMail"] = str(cols[10])

    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "skill_compen_config, end"


def abtest_config():
    """
    ab test配置
    """
    print "abtest_config, start, "
    outPath = getOutPath("abTest")
    wb = getWorkBook()
    ws = wb.get_sheet_by_name("ABTest")
    config = collections.OrderedDict()
    startRowNum = 4
    h = 0
    for row in ws.rows:
        h = h + 1
        if h < startRowNum:
            continue
        cols = []
        for cell in row:
            cols.append(cell.value)
        one = collections.OrderedDict()
        config["abcTest"] = one
        if cols[0]:
            one["mode"] = str(cols[0])
        if cols[1]:
            one["registerRechargeBonus"] = json.loads(cols[1])
        if cols[2]:
            one["firstCurve"] = json.loads(cols[2])
        if cols[3]:
            one["enterLimit"] = json.loads(cols[3])
        if cols[4]:
            one["waveChange"] = json.loads(cols[4])
        if cols[5]:
            one["enableNewbie8DaysGift"] = json.loads(cols[5])
        if cols[6]:
            one["enableNewbieLottery"] = json.loads(cols[6])
        if cols[7]:
            one["enableUserCoupon"] = json.loads(cols[7])
        if cols[8]:
            one["extraRechargeBonus"] = json.loads(cols[8])

    result = json.dumps(config, indent=4, ensure_ascii=False)
    outHandle = open(outPath, "w")
    outHandle.write(result)
    outHandle.close()
    print "abtest_config, end"


def getWorkBook(filename="newfish.xlsm"):
    configFile = os.path.split(os.path.realpath(__file__))[0] + "/%s" % filename
    return load_workbook(filename=configFile, read_only=True, data_only=True)


def getOutPath(dirname, filename="0.json"):
    global ServerPath
    dirPath = os.path.split(os.path.realpath(__file__))[0] + ServerPath + "/%s/" % dirname
    filePath = dirPath + filename
    if not os.path.exists(dirPath):
        os.makedirs(dirPath)
    return filePath


def process_single_config(idx, task_queue, cost_queue, err_queue, sp):
    """
    处理单个配置文件
    """
    reload(sys)
    sys.setdefaultencoding("utf-8")
    print "---- %d cpu start!" % idx
    global ServerPath
    ServerPath = sp
    while not task_queue.empty():
        confFunc, arg = task_queue.get()
        try:
            t1 = time.time()
            if arg is None:
                confFunc()
            else:
                confFunc(arg)
            _costTime = time.time() - t1
            if arg is None:
                cost_queue.put((confFunc.__name__, _costTime))
            else:
                cost_queue.put((confFunc.__name__ + "_" + str(arg), _costTime))
        except Exception, e:
            print "=========== export ", confFunc.__name__, " failed ! ", traceback.format_exc()
            err_queue.put((confFunc.__name__ + "_" + str(arg), e))


# 配置列表
config_list = [
    # (activity_config, None),
    # (multi_lang_text, None),
    (newbie7DaysGift_config, None),
    (lottery_ticket_config, None),
    (pass_card_config, None),
    (skill_compen_config, None),
    (grand_prize_pool_config, None),
    # (piggy_bank_config, None),
    (prizewheel_config, None),
    (grandPrixPrizewheel_config, None),
    # (time_limited_store_config, None),
    (levelRewards_config, None),
    (level_funds_config, None),
    (level_funds_config, 25794),
    (super_egg_config, None),
    (updateVerRewards_config, None),
    (bigPrize_config, None),
    (compAct_config, None),
    # (checkin_config, None),
    (fish_config, None),
    (weapon_config, None),
    (ulevel_config, None),
    (skill_grade_config, None),
    (skill_star_config, None),
    # (main_quest_config, None),
    # (daily_quest_config, None),
    (cmptt_task_config, None),
    (ncmptt_task_config, None),
    (bonus_task_config, None),
    (guide_task_config, None),
    (probability_config, None),
    (dynamic_odds_config, None),
    # (lottery_pool_config, None),
    (catch_drop_config, None),
    # (vip_config, None),
    # (fixed_multiple_fish, None),
    (call_multiple_fish, None),
    (fishBonus_config, None),
    # (match_multiple_fish, None),
    # (achievement_config, None),
    # (honor_config, None),
    (weaponPowerRate_config, None),
    (weapon_stage_count_config, None),
    (gunskin_config, None),
    # (plyerBuffer_config, None),
    # (randomMultipleFish_config, None),
    (gunLevel, None),
    (tableTask, None),
    (inviteTask_config, None),
    (rechargePool_config, None),
    # (share_config, None),
    (fly_pig_config, None),
    (starfish_roulette_config, None),
    (surpass_target_config, None),
    (special_item_config, None),
    (robot_config, None),
    (report_fish_cb_config, None),
    # (treasure_config, None),
    (statistic_prop_config, None),
    (slot_machine_config, None),
    (money_tree_config, None),
    (canned_fish_config, None),
    (credit_store_config, None),
    (supply_box_config, None),
    (festival_turntable_config, None),
    (collect_item_config, None),
    # (abtest_config, None),
    # (world_boss_config, None),
    (terror_fish, None),
]

if __name__ == "__main__":
    TestConfPath = ""
    RealeaseConfPath = ""
    print "begin"
    t1 = int(time.time())
    ServerPath = '/../../../../../../wx_superboss/trunk/xxfish_dev/config37/game/44'          # 默认练习路径
    if len(sys.argv) > 1 and sys.argv[1] == '-d':
        ServerPath = '/../../../../../../xxfish_test/config37/game/44'
    # else:
    #     ServerPath = "/../../../../../../../gameServer/hall37/newfish-py/wx_superboss/xxfish_dev/config37/game/44"

    import platform
    _system = platform.system()
    if _system == "Darwin" and len(sys.argv) > 1 and sys.argv[1] == "-k":
        upConfig = "svn up .%s" % ServerPath
        os.system("svn up ./")
        os.system(upConfig)
        print "update %s finish" % ServerPath

    print "----- start, config"
    conf_queue = Queue(len(config_list))
    for conf in config_list:
        conf_queue.put(conf)
    cost_queue = Queue()
    err_queue = Queue()
    isUseMultiProcess = len(sys.argv) > 2 and sys.argv[2] == "-m" and cpu_count() > 1
    if isUseMultiProcess:
        freeze_support()
        processList = [Process(target=process_single_config, args=(i + 1, conf_queue, cost_queue, err_queue, ServerPath)) for i in range(cpu_count())]
        for p in processList:
            p.start()
        for p in processList:
            p.join()
    else:
        process_single_config(0, conf_queue, cost_queue, err_queue, ServerPath)

    # s1 = set()
    # while not cost_queue.empty():
    #     fn, costTime = cost_queue.get()
    #     s1.add(fn)
    #     if costTime > 5:
    #         print fn, costTime
    while not err_queue.empty():
        print err_queue.get()
    # windows需要转换json的换行符.
    if _system == "Windows":
        # print "json format: windows->unix, start"
        import os
        if TestConfPath:
            os.system(r".\jsonDos2Unix.bat %s" % TestConfPath)
        if RealeaseConfPath:
            os.system(r".\jsonDos2Unix.bat %s" % RealeaseConfPath)
        # print "json format: windows->unix, end"

    t2 = int(time.time())
    print "----- load config successfully"
    print "----- out full path: %s" % os.path.abspath(os.path.split(os.path.realpath(__file__))[0] + ServerPath)
    if isUseMultiProcess:
        print "----- %d cpu, process %d files, export json cost %d s" % (cpu_count(), len(config_list), t2 - t1)
    else:
        print "----- process %d files, export json cost %d s" % (len(config_list), t2 - t1)

