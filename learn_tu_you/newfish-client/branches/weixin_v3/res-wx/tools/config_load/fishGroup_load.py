# -*- coding=utf-8 -*-
"""
Created by lichen on 17/3/28.
鱼阵随机生成脚本
从TracksNew目录中获取单鱼种子鱼阵，按照fishGroup.xlsx中的鱼阵规划，随机组合生成鱼阵文件并导出到fish_group目录下
"""

from openpyxl import load_workbook
import os, sys, json, copy, random
import collections
import platform


def fishGroup_config():
    br = "\r\n"
    if platform.system() == "Windows":
        br = "\n"
    configFile = os.path.split(os.path.realpath(__file__))[0] + "/fishGroup.xlsx"
    outPath = os.path.split(os.path.realpath(__file__))[0] + "/fish_group/0.json"
    
    wb = load_workbook(filename = configFile, read_only = True, data_only = True)
    wsFishGroup = wb.get_sheet_by_name("FishGroup")
    scenes = collections.OrderedDict()
    sceneRowNum = 2
    i = 0
    for row in wsFishGroup.rows:
        i = i + 1
        if i < sceneRowNum:
            continue
        sceneCols = []
        for cell in row:
            sceneCols.append(cell.value)
        if not sceneCols[0]:
            continue
        scene = collections.OrderedDict()
        scene["roomId"] = int(sceneCols[0])
        scene["totalNum"] = int(sceneCols[1])
        scene["minTime"] = float(sceneCols[2])
        scene["maxTime"] = float(sceneCols[3])
        scene["locMinTime"] = float(sceneCols[4])
        scene["locMaxTime"] = float(sceneCols[5])
        scene["totalTime"] = float(sceneCols[6])
        scenes[int(sceneCols[0])] = scene

    # result = json.dumps(scenes, indent=4)
    # outHandle = open(outPath, "w")   
    # outHandle.write(result)  
    # outHandle.close() 

    wb2 = load_workbook(filename = configFile, read_only = True, data_only = True)
    wsFishWeight = wb2.get_sheet_by_name("FishWeight")
    weights = collections.OrderedDict()
    i = 0
    weightRowNum = 2
    for key in scenes.keys():
        weights[str(key)] = []
    for row in wsFishWeight.rows:
        i = i + 1
        if i < weightRowNum:
            continue
        weightCols = []
        for cell in row:
            weightCols.append(cell.value)
        if not weightCols[0]:
            continue  
        x = 0
        for key in scenes.keys():
            if int(weightCols[x+3]):
                weight = collections.OrderedDict()
                weight["fishId"] = int(weightCols[0])
                weight["path"] = float(weightCols[1])
                weight["weight"] = int(weightCols[x+3])
                weights[str(key)].append(weight)
            x += 1

    # result = json.dumps(weights, indent=4)
    # outHandle = open(outPath, "w")   
    # outHandle.write(result)  
    # outHandle.close()

    groupsPath = os.path.split(os.path.realpath(__file__))[0] + "/TracksNew"
    fishJson = {}
    for fileName in os.listdir(groupsPath):
        abstractFilePath = os.path.join(groupsPath, fileName)
        fishJson[fileName] = {}
        fishJson[fileName]["id"] = fileName
        fishJson[fileName]["fishes"] = []
        fHandle = open(abstractFilePath, "r")
        lines = fHandle.readlines()
        for i in range(len(lines)):
            line = lines[i]
            if i == 0:
                fishJson[fileName]["totalTime"] = float(lines[i])
            else:
                line.strip()
                if len(line) == 0:
                    continue
                fields = line.split("|")
                # print fields[0]
                fishJson[fileName]["fishes"].append({
                    "fishType": int(fields[0]),
                    "enterTime": float(fields[1]),
                    "exitTime": float(fields[2]),
                    "path": int(fields[3])
                })
        fHandle.close()
        # print fishJson

    outFilePath = os.path.split(os.path.realpath(__file__))[0] + "/fish_group/"
    for roomId, scene in scenes.items():
        for i in xrange(scene["totalNum"]):
            totalTime = 0
            lastEnterTime = 0
            maxEnterTime = 0
            intervalTime = 0
            fishGroup = []
            outFileName = "group_%d_%d" % (roomId, i + 1)
            print outFileName
            while maxEnterTime < scene["totalTime"]:
                intervalTime += float("%.2f" % random.uniform(scene["minTime"], scene["maxTime"]))
                if maxEnterTime == 0:
                    intervalTime = 0
                fishWeightList = weights[str(roomId)]
                totalWeight = sum([int(weight["weight"]) for weight in fishWeightList])
                if totalWeight <= 0:
                    continue
                fishPath = 0
                for loc in range(4):
                    fishId = 0
                    randWeight = random.randint(0, totalWeight)
                    locTime = float("%.2f" % random.uniform(scene["locMinTime"], scene["locMaxTime"]))
                    for weight in fishWeightList:
                        if randWeight - weight["weight"] <= 0:
                            fishId = weight["fishId"]
                            fishPath = weight["path"]
                            break
                        randWeight -= weight["weight"]
                    path = random.randint(1, fishPath)
                    print fishJson["%d_%d_%d" % (fishId, path, loc + 1)]
                    fishes = copy.deepcopy(fishJson["%d_%d_%d" % (fishId, path, loc + 1)]["fishes"])
                    for fish in fishes:
                        enterTime = fish["enterTime"]
                        exitTime = fish["exitTime"]
                        fish["enterTime"] = float("%.2f" % (enterTime + intervalTime + locTime))
                        fish["exitTime"] = float("%.2f" % (exitTime + intervalTime + locTime))
                        if fish["enterTime"] > maxEnterTime:
                            maxEnterTime = fish["enterTime"]
                        if fish["exitTime"] > totalTime:
                            totalTime = fish["exitTime"]
                    if maxEnterTime >= scene["totalTime"]:
                        print str(lastEnterTime) + "|" + str(totalTime)
                        break
                    lastEnterTime = maxEnterTime
                    fishGroup.extend(fishes)
            count = 0
            fHandle = open(outFilePath + outFileName, "w+")
            for fishMap in fishGroup:
                count += 1
                if count == 1:
                    fHandle.write("%.2f%s" % (totalTime, br))
                    # print totalTime
                fHandle.write(str(fishMap["fishType"]) + "|" + "%.2f" % fishMap["enterTime"] + "|" + "%.2f" % fishMap["exitTime"] + "|" + str(fishMap["path"]) + br)
            fHandle.close


if __name__ == "__main__":
    print "begin"
    fishGroup_config()
    print "load fishGroup_config successfully"