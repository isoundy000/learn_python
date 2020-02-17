#!/usr/bin/env python
# -*- coding:utf-8 -*-
import os
import sys
import time
import openpyxl
import traceback

from lib.utils.debug import print_log, trackback
import settings as _settings
from models.config import ConfigManager
from models.config import ConfigVersionManager
# from apps.config import reload_all, update_config
from game_config import make_version

from admin.upload import xls_convert, label_check
from admin.upload import config_templates as CT         # 此处引入是为了做配置检查

###################################
##########
DEBUG = False
##########
###################################


def config_save(config_name, config_str):
    """存储方法
    """
    config = open(_settings.BASE_ROOT + '/config_database/' + config_name + '.txt', 'w')
    config.write(config_str.encode('utf-8'))
    config.close()
    # update_config()
    # reload_all()


def upload(file_obj, server_id='', config_type=1):
    '''
    上传文件
    '''
    """# upload_single: docstring
    args:
        re:    ---    arg
    returns:
        0    ---
    """
    # 获取当前时分秒
    file_name_part = '-'.join([str(i) for i in time.localtime()[:6]])

    # 保存文件
    file_dir = os.path.join(_settings.BASE_ROOT, 'upload_xls')
    if not os.path.exists(file_dir):
        os.makedirs(file_dir)

    file_name = os.path.join(file_dir, file_obj[0]['filename'].strip('.xlsx') + '_' + file_name_part + '.xlsx')
    filebody = file_obj[0]['body']
    hfile = open(file_name, 'wb+')
    hfile.write(filebody)
    hfile.close()

    return import_file(file_name, server_id, config_type=config_type)


def import_file(filename, server_id='', config_type=1):
    """导入一个xlsx文件配置
    args:
        filename: xlsx文件路径
    returns:
        []: 导入成功的配置名列表
    """
    done_list = []
    cv = ConfigVersionManager.get_config_version_obj(config_type=config_type)
    import game_config
    config_name_list_raw = game_config.config_name_list
    sub_config_name_list = dict(zip(
        [i[0] for i in config_name_list_raw if i[3]],
        [i[1] for i in config_name_list_raw if i[3]]
    ))

    try:
        data, done_list, need_private_city_test = trans_config(filename, cv, done_list, config_type=config_type)
    except Exception, e:
        etype, value, tb = sys.exc_info()
        line = traceback.format_exception_only(etype, value)
        line_str = '-'.join(line)
        return 100, line_str.replace('\\', '')

    # -99 用于配置表内容逻辑有误时的return code
    if done_list == 'bug':
        return -99, data

    for config_name, sheet_title, str_config in data:
        c = ConfigManager.get_config_obj(config_name, config_type=config_type)
        if DEBUG or c.value != str_config:
            m_version = make_version(str_config)
            c.value = str_config
            c.version = m_version
            c.save()
            cv.update_single(config_name, hex_version=m_version)
            sub_funcs = sub_config_name_list.get(config_name)
            for sub_func in sub_funcs:
                sub_func(config_name, sub_func.func_name, cv, force=True)

            print 'config: ', config_name, ' saved'
            done_list.append(sheet_title)

    cv.save()

    if need_private_city_test:
        from logics.private_city import city_map_test
        tb, local_dict = city_map_test()
        if tb:
            return 1, [tb, local_dict]
    return 0, done_list


def trans_config(filename, cv=None, done_list=None, config_type=1):
    """通过对应的数据表 生成相应的配置"""
    rs = []

    xl = openpyxl.load_workbook(filename=filename)      # 仅支持xlsx   use_iterators=True
    import game_config
    config_name_list_raw = game_config.config_name_list
    config_name_list = zip(
        [i[0] for i in config_name_list_raw if i[3]],
        [i[4] for i in config_name_list_raw if i[3]]
    )

    done_list = done_list or []
    cv = cv or ConfigVersionManager.get_config_version_obj(config_type=config_type)
    need_private_city_test = False
    for sheet in xl.worksheets:
        config_name = ''
        sheet_title = sheet.title.strip()
        if 'map' in sheet_title or 'title_detail' in sheet_title:
            need_private_city_test = True

        # 特殊处理
        tt = sheet_title.split('_')
        if len(tt) == 2 and tt[1].isdigit():
            if tt[0] == 'map':
                upload_map_status = upload_map(tt[1], sheet, config_type=config_type)
                if upload_map_status:
                    cv.update_single('map', hex_version=upload_map_status)
                continue
            elif tt[0] == 'middle':
                upload_middle_map_data_status = upload_middle_map_data(tt[1], sheet, config_type=config_type)
                if upload_middle_map_data_status:
                    cv.update_single('middle_map_data', hex_version=upload_middle_map_data_status)
                    done_list.append(sheet_title)
                continue

        # 处理box_reward_
        if len(tt) == 3 and '%s_%s' % (tt[0], tt[1]) == 'box_reward':
            upload_box_reward_status = upload_box_reward(int(tt[2]), sheet, config_type=config_type)
            if upload_box_reward_status:
                cv.update_single('box_reward_new', hex_version=upload_box_reward_status)
                done_list.append(sheet_title)
            continue

        # 特殊处理
        for i in config_name_list:
            if sheet_title in i:
                config_name = i[0]
        if not config_name:
            continue

        if config_name == 'guide':
            upload_guide_status = upload_guide(config_name, sheet, config_type=config_type)
            if upload_guide_status:
                cv.update_single(config_name, hex_version=upload_guide_status)
                done_list.append(sheet_title)
            continue

        data = xls_convert.to_pyobj(sheet)
        # 针对在config_templates里指明的label进行内容逻辑测试
        # 含'guide', 'middle', 'map', 'box_reward'的表在获得本函数使用的data前，已进行预处理。
        # 如要对上述表进行检测，需阅读预处理代码，避免潜在冲突
        has_bug = content_logic_check(config_name, sheet_title, data)
        if has_bug:
            return has_bug, 'bug', need_private_city_test

        try:
            str_config = xls_convert.to_config_string(config_name, data)
        except Exception, e:
            etype, value, tb = sys.exc_info()
            line = traceback.format_exception_only(etype, value)
            raise KeyError("table=", sheet_title, line)

        try:
            d = eval(str_config)
        except Exception, e:
            file_dir = os.path.join(_settings.BASE_ROOT, 'logs')
            if not os.path.exists(file_dir):
                os.makedirs(file_dir)
            error_file = os.path.join(file_dir, 'config_error.py')
            with open(error_file, 'w') as f:
                f.write('# encoding: utf-8 \n')
                f.write('# config_name: %s \n' % config_name)
                f.write(str_config.encode('utf-8'))
            traceback()
            raise e
        rs.append((config_name, sheet_title, str_config))
    return rs, done_list, need_private_city_test


def upload_map(map_id, sheet, config_type=1):
    """# upload_single: docstring
    args:
        re:    ---    arg
    returns:
        0    ---
    """
    result = []
    iter_r = sheet.iter_rows()
    for i, row in enumerate(iter_r):
        if i == 0: continue
        if row[0].internal_value is None:
            break
        r = []
        for cell in row:
            if cell.internal_value is None:
                break
            r.append(int(cell.internal_value))
        result.append(r)

    c = ConfigManager.get_config_obj('map', config_type=config_type)
    if DEBUG or c.value.get(str(map_id)) != result:
        data = [(k, v) for k, v in c.value.iteritems()]
        data.sort(key=lambda x: x[0])
        m_version = make_version(data)
        c.value[str(map_id)] = result
        c.version = m_version
        c.save()
        return m_version


def upload_middle_map_data(map_id, sheet, config_type=1):
    """# upload_middle_map_data: docstring
    args:
        map_id, sheet:    ---    arg
    returns:
        0    ---
    """
    result = []
    iter_r = sheet.iter_rows()
    for i, row in enumerate(iter_r):
        if i == 0: continue
        if row[0].internal_value is None:
            break
        r = []
        for cell in row:
            v = cell.internal_value
            if v is None:
                break
            try:
                v = int(v)
                r.append(xls_convert.mapping['int_list']("""[%d, 1]""" % v))
            except ValueError:
                r.append(xls_convert.mapping['int_list'](v))

        result.append(r)

    c = ConfigManager.get_config_obj('middle_map_data', config_type=config_type)
    if DEBUG or c.value.get(str(map_id)) != result:
        data = [(k, v) for k, v in c.value.iteritems()]
        data.sort(key=lambda x: x[0])
        m_version = make_version(data)
        c.value[str(map_id)] = result
        c.version = m_version
        c.save()
        return m_version


def upload_box_reward(box_id, sheet, config_type=1):
    """# upload_single: docstring
    args:
        re:    ---    arg
    returns:
        0    ---
    """
    data = xls_convert.to_pyobj(sheet)

    str_config = xls_convert.to_config_string('box_reward', data)
    dict_config = eval(str_config)

    c = ConfigManager.get_config_obj('box_reward_new', config_type=config_type)
    if DEBUG or c.value.get(box_id) != dict_config:
        data = [(k, sorted(v.iteritems(), key=lambda x: x[1])) for k, v in c.value.iteritems()]
        data.sort(key=lambda x: x[0])
        m_version = make_version(data)
        c.value[box_id] = dict_config
        c.version = m_version
        c.save()
        return m_version


def upload_guide(config_name, sheet, config_type=1):
    """分组新手引导步骤
    args:
        config_name: 配置名字
        sheet: xlsx sheet对象
    """
    data = xls_convert.to_pyobj(sheet)

    str_config = xls_convert.to_config_string(config_name, data)
    dict_config = eval(str_config)

    result = {}
    for guide_id, obj in dict_config.iteritems():
        id_data = result.setdefault(obj['guide_team'], {})
        id_data[guide_id] = obj

    c = ConfigManager.get_config_obj(config_name, config_type=config_type)
    if DEBUG or c.value != result:
        m_version = make_version(str_config)
        c.value = result
        c.version = m_version
        c.save()
        return m_version


def content_logic_check(config_name, sheet_title, data):
    """当配置内容有逻辑错误时，返回错误提示信息，否则返回None。"""
    template_tuple = getattr(CT, config_name)()
    check_func_list = []
    if len(template_tuple) == 2:
        trans, funcs = template_tuple
    elif len(template_tuple) == 3:
        trans, funcs, check_func_list = template_tuple
    else:
        raise TypeError("to_config_string takes 2 or 3 arguments, (%d given)" % len(template_tuple))
    for check_func_name in check_func_list:
        check_func = getattr(label_check, check_func_name)
        has_bug = check_func(sheet_title, data)
        if has_bug:
            return has_bug


def static_import(dirpath, config_type=1):
    """导入目录下所有xlsx配置
    args:
        dirpath: 目录路径
    """
    error_files = []
    for filename in os.listdir(dirpath):
        if 'xls' not in filename:
            continue
        filepath = os.path.join(dirpath, filename)
        try:
            import_file(filepath, config_type=config_type)
        except:
            error_files.append(filename)

    if error_files:
        print 'error_files:', error_files


if __name__ == '__main__':
    filename = r'E:\chuangshi_xlsx\reward.xlsx'
    # print import_file(filename)
    static_import(r'e:\chuangshi_xlsx', 1)