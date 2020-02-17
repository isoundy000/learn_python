#!/usr/bin/env python
# -*- coding:utf-8 -*-

# openpyxl中读大批量数据的方法——Optimized reader
# 有时候，我们需要打开或写数据量非常大的XLSX文件，而openpyxl的通用方法将无法处理如此大的负载。令人高兴的是，
# openpyxl有两种模式，使我们可以用常量级的内存消耗来读取和写入无限量的数据。
# 本文首先介绍读大量数据的方法，写的方法我没有尝试，以后有了会贴上来，需要的可以参考http://packages.python.org/openpyxl/optimized.html。
# 首先我们需要打开一个excel表格，与以往的打开方式少有不同，比如我们打开一个文件名为haggle的xlsx文件，其调用方式应该为
# wb = load_workbook(filename = 'haggle.xlsx',use_iterators=True)我们发现后面多了一个use_iterators=True。
# 其次，我们需要打开该工作薄的一个具体的工作表（sheet）：ws=wb.get_sheet_by_name(name='Sheet1')此时ws便是一个IterableWorksheet
# 接下来我们便可以对该工作表ws进行处理了，我们需要着重解释一下代码中的一个函数，先贴上代码。该代码的功能是统计C列1-21282行分别出现0,1到35的次数，并将其写到另外一个excel中。

from openpyxl import load_workbook
from openpyxl import Workbook

dest_filename = 'haggle.xlsx'
new_filename = 'result_sf.xlsx'
wb = load_workbook(filename=dest_filename, read_only=True)
# keep_vba和data_only该两个参数均为布尔参数，将其设置为True，即打开excel，开启并保留宏，excel中工作簿中包含公式的单元格，需读取公式结果时，则显示公式的计算结果
# 附上load_workbook()6个参数：
#  - filename: string类型，文件路径 或路径对象
#  - read_only：布尔型，超大型文件，为节省内存，建议开启此参数
#  - keep_vba ：布尔型，True则保留vba代码
#  - guess_types：布尔型， 读取单元格数据类型时，启用或禁用类型推断
#  - data_only：布尔型，True则包含公式的单元格，显示最近计算结果
#  - keep_links：布尔型，True保留外部链接
wb1 = Workbook()
ws = wb.get_sheet_by_name("Sheet4")
ws1 = wb1.create_sheet(0, "socialfrequency")
for i in range(0, 36):
    count = 0
    for row in ws.iter_rows('C1:D21282'):  # 遍历行
        for cell in row:  # 对每行遍历每个数据单元
            if i == cell.internal_value:
                count = count + 1
            else:
                continue
    ws1.cell(row=i, column=7).value = i
    ws1.cell(row=i, column=8).value = count

wb1.save(new_filename)

# 我们需要着重解释的代码为ws.iter_rows('C1:D21282')该函数返回的是C列的1到21282行，该函数居然包括C不包括D但是却又包括1和21282行，差点搞死我。
# 注意：如果想正常运行程序，需要大家讲程序中的所有注释汉字去掉，这涉及到编码的问题，目前我还没有完全搞明白的东西。