#!/usr/bin/env python
# -*- coding:utf-8 -*-

from openpyxl import Workbook
# 创建workbook:
wb = Workbook()
print wb

# 创建sheet:
# 1,默认表，在创建wb时默认就创建了sheet，使用
ws = wb.active    # 得到
print ws
# 2,创建新表
ws1 = wb.create_sheet()     # 插入到结尾
ws2 = wb.create_sheet(index=0)    # 插入到开头
n = 1000
wsn = wb.create_sheet(index=n)    # 插入到第n个位置
# 创建新行:
# 这个方法openpyxl没有给定函数，overstackflow有相关函数，请见：[添加行][3]
# 创建新列
# 这个方法openpyxl没有给定函数，overstackflow有相关函数，请见：[添加列][4]
# 保存
# wb.save(file)

# 读
#     读workbook 文档
# 使用:
from openpyxl import load_workbook
filename = 'fuck.xls'
wb = load_workbook(filename)
# 来源于：
# openpyxl.reader.excel.load_workbook(filename, read_only=False, use_iterators=False, keep_vba=False, guess_types=False, data_only=False)
# 注：
# filename是路径+文件名*
# read_only:只读模式，会使得读取更快
# use_iterators:该lazy模式，默认采用只读模式，所有的worksheet将openpyxl.worksheet.iter_worksheet.IterableWorksheet类型的
# 它的特点是跳过empty cells


# 读worksheet
# ws = wb[sheetname]
# ws = wb.get_sheet_by_name(sheetname)
# 想要知道sheet名可以用
# wb.get_sheet_names()得到

# 逐行读
ws.iter_rows(range_string=None, row_offset=0, column_offset=0)
# range-string(string)-单元格的范围：例如('A1:C4')
# row_offset-添加行
# column_offset-添加列
# 返回一个生成器,
# 注意取值时要用value,例如：
for var in ws.iter_rows():
    print var[0].value

# 读指定行、指定列
rows=ws.rows    # row是可迭代的
columns=ws.columns  # column是可迭代的
# 打印第n行所有数据
print rows[n]   # 不需要用.value
print columns[n]    # 不需要用.value

# 读连续单元格
cell_range = ws['A1':'C2']
# 读指定的单元格
ws.cell('B12').value
ws.cell(row=12, column=2).value

# 读所有单元格数据
ws.get_cell_collection()

# 写一行源代码和解释
# ws.append(iterable)

# 添加一行到当前sheet的最底部
# iterable必须是list,tuple,dict,range,generator类型的。
# 1,如果是list,将list从头到尾顺序添加。
# 2，如果是dict,按照相应的键添加相应的键值。
# append([‘This is A1’, ‘This is B1’, ‘This is C1’])
# append({‘A’ : ‘This is A1’, ‘C’ : ‘This is C1’})
# append({1 : ‘This is A1’, 3 : ‘This is C1’})

# 写一列
# 写一格
ws['A4'] = 4