#!/usr/bin/env python
# -*- coding:utf-8 -*-

__author__ = 'ghou'


# from openpyxl import load_workbook
# 打开已有
# wb = load_workbook("template.xlsx")
# print(wb.sheetnames)    # ['Sheet1', 'Sheet2', 'Sheet3']
#
# sheet = wb.get_sheet_by_name("Sheet3")
# print(sheet["C"])       # (<Cell Sheet3.C1>, <Cell Sheet3.C2>, <Cell Sheet3.C3>, <Cell Sheet3.C4>, <Cell Sheet3.C5>, <Cell Sheet3.C6>, <Cell Sheet3.C7>, <Cell Sheet3.C8>, <Cell Sheet3.C9>, <Cell Sheet3.C10>)      <-第C列
# print(sheet["4"])       # (<Cell Sheet3.A4>, <Cell Sheet3.B4>, <Cell Sheet3.C4>, <Cell Sheet3.D4>, <Cell Sheet3.E4>)     <-第4行
# print(sheet["C4"].value)    # c4     <-第C4格的值
# print(sheet.max_row)    # 10     <-最大行数
# print(sheet.max_column)    # 5     <-最大列数
# for i in sheet["C"]:
#     print(i.value)    # c1 c2 c3 c4 c5 c6 c7 c8 c9 c10     <-C列中的所有值


# 3、写入xlsx
import datetime
from openpyxl import Workbook
# 实例化
wb = Workbook()
# 激活worksheet
sheet = wb.active
# 找到活动的sheet页。空的excel表默认的sheet页就叫Sheet，如果想改名字，可以直接给title属性赋值。
sheet.title = "Sheet1"
# 这个属性是可读可写的。当然，这个只针对当前活动页，别的页的话，可以用create_sheet和remove_sheet进行添加和删除。
# 往sheet页里面写内容就比较简单了，跟上面读一样，
# 方式一:数据可以直接分配到单元格中(可以输入公式) ws['A1'] = 42 # 方式二:可
# sheet['C3'] = 'Hello world!'
# 方式二：可以附加行，从第一列开始附加(从最下方空白处，最左开始)(可以输入多行)
sheet.append([1, 2, 3])
# 方式三：Python 类型会被自动转换
sheet['A3'] = datetime.datetime.now().strftime("%Y-%m-%d")
for i in range(10):
    sheet["A%d" % (i+1)].value = i + 1
    sheet["B%d" % (i+1)].value = i + 1
    sheet["C%d" % (i+1)].value = i + 1
# 我们还可以进行花式操作，比如写写公式：
# sheet["E1"].value = "=SUM(A:A)"
# 最后记得保存
wb.save('template.xlsx')


# 4. 创建表（sheet）
# 方式一：插入到最后(default)
ws1 = wb.create_sheet("Mysheet")
# 方式二：插入到最开始的位置
ws2 = wb.create_sheet("Mysheet", 0)

# 5. 选择表（sheet）
# sheet 名称可以作为 key 进行索引
ws3 = wb["New Title"]
ws4 = wb.get_sheet_by_name("New Title")
print ws3 is ws4

# 6. 查看表名（sheet）
# 显示所有表名
print(wb.sheetnames)
# ['Sheet2', 'New Title', 'Sheet1']
# 遍历所有表
for sheet in wb:
    print(sheet.title)

# 7. 访问单元格（call）
# ① 单一单元格访问
# 方法一
# c = ws['A4']
# 方法二：row 行；column 列
# d = ws.cell(row=4, column=2, value=10)
# 方法三：只要访问就创建
# for i in range(1, 101):
#   for j in range(1, 101):
#       ws.cell(row=i, column=j)

# ② 多单元格访问
# 通过切片
# >>> cell_range = ws['A1':'C2']
# # 通过行(列)
# >>> colC = ws['C']
# >>> col_range = ws['C:D']
# >>> row10 = ws[10]
# >>> row_range = ws[5:10]
# # 通过指定范围(行 → 行)
# >>> for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
# ...  for cell in row:
# ...    print(cell)
# <Cell Sheet1.A1>
# <Cell Sheet1.B1>
# <Cell Sheet1.C1>
# <Cell Sheet1.A2>
# <Cell Sheet1.B2>
# <Cell Sheet1.C2>　
# # 通过指定范围(列 → 列)
# >>> for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
# ...  for cell in row:
# ...    print(cell)
# <Cell Sheet1.A1>
# <Cell Sheet1.B1>
# <Cell Sheet1.C1>
# <Cell Sheet1.A2>
# <Cell Sheet1.B2>
# <Cell Sheet1.C2>
# # 遍历所有 方法一
# >>> ws = wb.active
# >>> ws['C9'] = 'hello world'
# >>> tuple(ws.rows)
# ((<Cell Sheet.A1>, <Cell Sheet.B1>, <Cell Sheet.C1>),
# (<Cell Sheet.A2>, <Cell Sheet.B2>, <Cell Sheet.C2>),
# ...
# (<Cell Sheet.A8>, <Cell Sheet.B8>, <Cell Sheet.C8>),
# (<Cell Sheet.A9>, <Cell Sheet.B9>, <Cell Sheet.C9>))
# # 遍历所有 方法二
# >>> tuple(ws.columns)
# ((<Cell Sheet.A1>,
# <Cell Sheet.A2>,
# <Cell Sheet.A3>,
# ...
# <Cell Sheet.B7>,
# <Cell Sheet.B8>,
# <Cell Sheet.B9>),
# (<Cell Sheet.C1>,
# ...
# <Cell Sheet.C8>,
# <Cell Sheet.C9>))
# 8. 保存数据
# wb.save('文件名称.xlsx')
# 9. 其他
# ① 改变 sheet 标签按钮颜色
# ws.sheet_properties.tabColor = "1072BA"
# ② 获取最大行，最大列
# 获得最大列和最大行
print(sheet.max_row)
print(sheet.max_column)

# ③ 获取每一行，每一列
# sheet.rows为生成器, 里面是每一行的数据，每一行又由一个tuple包裹。
# sheet.columns类似，不过里面是每个tuple是每一列的单元格。
# 因为按行，所以返回A1, B1, C1这样的顺序
for row in sheet.rows:
    for cell in row:
        print(cell.value)

# A1, A2, A3这样的顺序
for column in sheet.columns:
    for cell in column:
        print(cell.value)

# ④ 根据数字得到字母，根据字母得到数字
from openpyxl.utils import get_column_letter, column_index_from_string

# 根据列的数字返回字母
print(get_column_letter(2))  # B
# 根据字母返回列的数字
print(column_index_from_string('D'))  # 4

# ⑤ 删除工作表
# 方式一
wb.remove(sheet)
# 方式二
del wb[sheet]

# ⑥ 矩阵置换（行 → 列）
rows = [
    ['Number', 'data1', 'data2'],
    [2, 40, 30],
    [3, 40, 25],
    [4, 50, 30],
    [5, 30, 10],
    [6, 25, 5],
    [7, 50, 10]
]

print list(zip(*rows))

# 注意 方法会舍弃缺少数据的列(行)
rows = [
  ['Number', 'data1', 'data2'],
  [2, 40  ], # 这里少一个数据
  [3, 40, 25],
  [4, 50, 30],
  [5, 30, 10],
  [6, 25, 5],
  [7, 50, 10],
]


print list(zip(*rows))

# 10. 设置单元格风格
# ① 需要导入的类
from openpyxl.styles import Font, colors, Alignment
# ② 字体
# 下面的代码指定了等线24号，加粗斜体，字体颜色红色。直接使用cell的font属性，将Font对象赋值给它。
bold_itatic_24_font = Font(name='等线', size=24, italic=True, color=colors.RED, bold=True)
sheet['A1'].font = bold_itatic_24_font
# ③ 对齐方式
# 也是直接使用cell的属性aligment，这里指定垂直居中和水平居中。除了center，还可以使用right、left等等参数。
# 设置B1中的数据垂直居中和水平居中
sheet['B1'].alignment = Alignment(horizontal='center', vertical='center')
# ④ 设置行高和列宽
# 第2行行高
sheet.row_dimensions[2].height = 40
# C列列宽
sheet.column_dimensions['C'].width = 30

# ⑤ 合并和拆分单元格
#     所谓合并单元格，即以合并区域的左上角的那个单元格为基准，覆盖其他单元格使之称为一个大的单元格。
#     相反，拆分单元格后将这个大单元格的值返回到原来的左上角位置。
# 合并单元格， 往左上角写入数据即可
sheet.merge_cells('B1:G1')  # 合并一行中的几个单元格
sheet.merge_cells('A1:C3')  # 合并一个矩形区域中的单元格

# 合并后只可以往左上角写入数据，也就是区间中:左边的坐标。
# 如果这些要合并的单元格都有数据，只会保留左上角的数据，其他则丢弃。换句话说若合并前不是在左上角写入数据，合并后单元格中不会有数据。
# 以下是拆分单元格的代码。拆分后，值回到A1位置。
sheet.unmerge_cells('A1:C3')

# 最后举个例子
import datetime
from random import choice
from time import time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 设置文件 mingc
addr = "openpyxl.xlsx"
# 打开文件
wb = load_workbook(addr)
# 创建一张新表
ws = wb.create_sheet()
# 第一行输入
ws.append(['TIME', 'TITLE', 'A-Z'])

# 输入内容（500行数据）
for i in range(500):
    TIME = datetime.datetime.now().strftime("%H:%M:%S")
    TITLE = str(time())
    A_Z = get_column_letter(choice(range(1, 50)))
    ws.append([TIME, TITLE, A_Z])

# 获取最大行
row_max = ws.max_row
# 获取最大列
con_max = ws.max_column
# 把上面写入内容打印在控制台
for j in ws.rows:  # we.rows 获取每一行数据
    for n in j:
        print(n.value)  # n.value 获取单元格的值
        print()
    # 保存，save（必须要写文件名（绝对地址）默认 py 同级目录下，只支持 xlsx 格式）
    wb.save(addr)